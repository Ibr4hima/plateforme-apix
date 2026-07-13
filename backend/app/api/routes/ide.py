from typing import Optional
from fastapi import APIRouter, Depends, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, distinct

from app.core.database import get_db
from app.core.auth import require_admin
from app.models.ide import IdeCnuced, IdeAnalyse, IdeKpiConfig
from app.models.shared import IdeCnucedMonde, RefGroupement, RefPays

router = APIRouter(prefix="/ide", tags=["IDE"])


# ── Calcul de tous les KPIs ───────────────────────────────────────────────────
def calc_kpis(rows: list) -> dict:
    def serie(direction, indicateur):
        return sorted([r for r in rows if r.direction==direction and r.indicateur==indicateur], key=lambda r: r.annee)
    def val(r): return float(r.valeur) if r.valeur is not None else None
    def last_valid(s): return next((r for r in reversed(s) if r.valeur is not None), None)
    def variation(v1, v2): return round((v1-v2)/abs(v2)*100,1) if v2 and v2!=0 else None

    fe = serie("entrant","flux"); fs = serie("sortant","flux")
    se = serie("entrant","stock"); ss = serie("sortant","stock")

    lfe = last_valid(fe); lfs = last_valid(fs)
    lse = last_valid(se); lss = last_valid(ss)

    pfe = next((r for r in reversed(fe) if r.annee < lfe.annee and r.valeur is not None), None) if lfe else None
    pfs = next((r for r in reversed(fs) if r.annee < lfs.annee and r.valeur is not None), None) if lfs else None
    pse = next((r for r in reversed(se) if r.annee < lse.annee and r.valeur is not None), None) if lse else None
    pss = next((r for r in reversed(ss) if r.annee < lss.annee and r.valeur is not None), None) if lss else None

    rec_fe = max((r for r in fe if r.valeur is not None), key=lambda r: r.valeur, default=None)
    rec_fs = max((r for r in fs if r.valeur is not None), key=lambda r: r.valeur, default=None)

    fe5   = [val(r) for r in fe[-5:] if r.valeur is not None]
    moy5  = round(sum(fe5)/len(fe5),1) if fe5 else None

    fe_now = val(lfe) if lfe else None
    fe_10  = next((val(r) for r in reversed(fe) if lfe and r.annee==lfe.annee-10 and r.valeur is not None), None)
    var10  = variation(fe_now, fe_10) if fe_now and fe_10 else None

    se_first = next((r for r in se if r.valeur is not None), None)
    tcam = None
    if se_first and lse and lse.annee > se_first.annee and se_first.valeur and lse.valeur:
        n = lse.annee - se_first.annee
        tcam = round(((float(lse.valeur)/float(se_first.valeur))**(1/n)-1)*100,1)

    se_1990 = next((val(r) for r in se if r.annee==1990 and r.valeur is not None), None)
    mult = round(val(lse)/se_1990,1) if lse and se_1990 and se_1990>0 else None

    bal  = round(val(lfe)-val(lfs),1) if lfe and lfs and lfe.annee==lfs.annee else None
    pbal_fe = next((r for r in reversed(fe) if lfe and r.annee==lfe.annee-1 and r.valeur is not None), None)
    pbal_fs = next((r for r in reversed(fs) if lfs and r.annee==lfs.annee-1 and r.valeur is not None), None)
    pbal = round(val(pbal_fe)-val(pbal_fs),1) if pbal_fe and pbal_fs else None

    ratio = round(val(lse)/val(lss),1) if lse and lss and val(lss) and val(lss)>0 else None

    consec = 0
    for i in range(len(fe)-1, 0, -1):
        if fe[i].valeur is not None and fe[i-1].valeur is not None and float(fe[i].valeur)>float(fe[i-1].valeur):
            consec += 1
        else:
            break

    total = round(sum(val(r) for r in fe if r.valeur is not None), 1)

    periode_rec  = [val(r) for r in fe if 2020<=r.annee<=2024 and r.valeur is not None]
    periode_prev = [val(r) for r in fe if 2015<=r.annee<=2019 and r.valeur is not None]
    moy_rec  = sum(periode_rec)/len(periode_rec)   if periode_rec  else None
    moy_prev = sum(periode_prev)/len(periode_prev) if periode_prev else None
    acceleration = variation(moy_rec, moy_prev) if moy_rec and moy_prev else None

    return {
        "flux_entrant_dernier":        {"valeur":val(lfe),"annee":lfe.annee if lfe else None,"variation":variation(val(lfe),val(pfe)),"unite":"M$","sens":"hausse_bien"},
        "flux_entrant_record":         {"valeur":val(rec_fe),"annee":rec_fe.annee if rec_fe else None,"variation":None,"unite":"M$","sens":"info"},
        "flux_entrant_moy5":           {"valeur":moy5,"annee":None,"variation":None,"unite":"M$","sens":"info"},
        "flux_entrant_var10":          {"valeur":var10,"annee":None,"variation":None,"unite":"%","sens":"hausse_bien"},
        "stock_entrant_dernier":       {"valeur":val(lse),"annee":lse.annee if lse else None,"variation":variation(val(lse),val(pse)),"unite":"M$","sens":"hausse_bien"},
        "stock_entrant_multiplication":{"valeur":mult,"annee":None,"variation":None,"unite":"×","sens":"info"},
        "stock_entrant_tcam":          {"valeur":tcam,"annee":None,"variation":None,"unite":"%","sens":"info"},
        "flux_sortant_dernier":        {"valeur":val(lfs),"annee":lfs.annee if lfs else None,"variation":variation(val(lfs),val(pfs)),"unite":"M$","sens":"baisse_bien"},
        "flux_sortant_record":         {"valeur":val(rec_fs),"annee":rec_fs.annee if rec_fs else None,"variation":None,"unite":"M$","sens":"info"},
        "stock_sortant_dernier":       {"valeur":val(lss),"annee":lss.annee if lss else None,"variation":variation(val(lss),val(pss)),"unite":"M$","sens":"baisse_bien"},
        "balance_derniere":            {"valeur":bal,"annee":lfe.annee if lfe else None,"variation":variation(bal,pbal) if bal and pbal else None,"unite":"M$","sens":"hausse_bien"},
        "ratio_stock":                 {"valeur":ratio,"annee":None,"variation":None,"unite":"×","sens":"info"},
        "annees_hausse_consecutive":   {"valeur":consec,"annee":None,"variation":None,"unite":"ans","sens":"info"},
        "flux_entrant_total_periode":  {"valeur":total,"annee":None,"variation":None,"unite":"M$","sens":"info"},
        "flux_entrant_acceleration":   {"valeur":acceleration,"annee":None,"variation":None,"unite":"%","sens":"hausse_bien"},
    }


# ── GET /ide/monde/groupements ────────────────────────────────────────────────
@router.get("/monde/groupements")
async def liste_monde_groupements(db: AsyncSession = Depends(get_db)):
    res = await db.execute(
        select(IdeCnucedMonde.code, IdeCnucedMonde.nom_fr, IdeCnucedMonde.categorie)
        .distinct()
        .order_by(IdeCnucedMonde.categorie, IdeCnucedMonde.nom_fr)
    )
    return [{"code": code, "nom_fr": nom_fr, "categorie": categorie}
            for code, nom_fr, categorie in res.all()]


# ── GET /ide/monde/details ────────────────────────────────────────────────────
# Données individuelles par pays pour un groupement donné (heatmap, radar, etc.)
@router.get("/monde/details")
async def get_monde_details(
    code:       str            = Query(...),
    annee_min:  Optional[int]  = Query(None),
    annee_max:  Optional[int]  = Query(None),
    indicateur: Optional[str]  = Query(None),
    direction:  Optional[str]  = Query(None),
    db: AsyncSession = Depends(get_db),
):
    grp_res = await db.execute(select(RefGroupement).where(RefGroupement.code == code))
    grp = grp_res.scalar_one_or_none()
    if not grp or not grp.pays_ids:
        return []

    q = select(IdeCnuced).where(IdeCnuced.ref_pays_id.in_(grp.pays_ids))
    if annee_min:   q = q.where(IdeCnuced.annee >= annee_min)
    if annee_max:   q = q.where(IdeCnuced.annee <= annee_max)
    if indicateur:  q = q.where(IdeCnuced.indicateur == indicateur)
    if direction:   q = q.where(IdeCnuced.direction == direction)
    q = q.order_by(IdeCnuced.pays, IdeCnuced.annee)

    res = await db.execute(q)
    rows = res.scalars().all()

    # Enrichir avec code_iso3
    ref_ids = list({r.ref_pays_id for r in rows if r.ref_pays_id})
    pays_map: dict = {}
    if ref_ids:
        rp_res = await db.execute(select(RefPays).where(RefPays.id.in_(ref_ids)))
        pays_map = {p.id: p for p in rp_res.scalars().all()}

    def f(v): return float(v) if v is not None else None
    return [{"pays": r.pays, "ref_pays_id": r.ref_pays_id,
             "code_iso3": pays_map.get(r.ref_pays_id, None) and pays_map[r.ref_pays_id].code_iso3,
             "annee": r.annee, "indicateur": r.indicateur, "direction": r.direction, "valeur": f(r.valeur)}
            for r in rows]


# ── GET /ide/monde ─────────────────────────────────────────────────────────────
@router.get("/monde")
async def get_monde(
    codes_list: Optional[str] = Query(None),   # "CEDEAO,G7"
    annee_min:  Optional[int] = Query(None),
    annee_max:  Optional[int] = Query(None),
    annees:     Optional[str] = Query(None),    # "2003,2008,2020"
    db: AsyncSession = Depends(get_db),
):
    q = select(IdeCnucedMonde)
    if codes_list:
        codes = [c.strip() for c in codes_list.split(",") if c.strip()]
        q = q.where(IdeCnucedMonde.code.in_(codes))
    if annees:
        liste = [int(a.strip()) for a in annees.split(",") if a.strip().isdigit()]
        if liste: q = q.where(IdeCnucedMonde.annee.in_(liste))
    else:
        if annee_min: q = q.where(IdeCnucedMonde.annee >= annee_min)
        if annee_max: q = q.where(IdeCnucedMonde.annee <= annee_max)
    q = q.order_by(IdeCnucedMonde.code, IdeCnucedMonde.annee)
    res = await db.execute(q)
    def f(v): return float(v) if v is not None else None
    return [{"code":r.code,"nom_fr":r.nom_fr,"categorie":r.categorie,"annee":r.annee,
             "indicateur":r.indicateur,"direction":r.direction,"moyenne":f(r.moyenne),
             "somme":f(r.somme),"min":f(r.min),"max":f(r.max),
             "variance":f(r.variance),"ecart_type":f(r.ecart_type)}
            for r in res.scalars().all()]


# ── GET /ide/cnuced — données brutes ──────────────────────────────────────────
@router.get("/cnuced")
async def get_cnuced(
    direction:  Optional[str] = Query(None),
    indicateur: Optional[str] = Query(None),
    pays:       Optional[str] = Query(None),       # un seul pays
    pays_list:  Optional[str] = Query(None),       # plusieurs pays séparés par virgule
    annee_min:  Optional[int] = Query(None),
    annee_max:  Optional[int] = Query(None),
    annees:     Optional[str] = Query(None),       # années spécifiques : "2003,2008,2020"
    db: AsyncSession = Depends(get_db),
):
    from sqlalchemy import and_, or_
    q = select(IdeCnuced)

    # Filtre pays : liste ou pays unique
    if pays_list:
        noms = [p.strip() for p in pays_list.split(",") if p.strip()]
        q = q.where(IdeCnuced.pays.in_(noms))
    elif pays:
        q = q.where(IdeCnuced.pays == pays)
    else:
        q = q.where(IdeCnuced.pays == "Sénégal")

    if direction:  q = q.where(IdeCnuced.direction  == direction)
    if indicateur: q = q.where(IdeCnuced.indicateur == indicateur)

    # Filtre années : spécifiques ou plage
    if annees:
        liste_annees = [int(a.strip()) for a in annees.split(",") if a.strip().isdigit()]
        if liste_annees: q = q.where(IdeCnuced.annee.in_(liste_annees))
    else:
        if annee_min: q = q.where(IdeCnuced.annee >= annee_min)
        if annee_max: q = q.where(IdeCnuced.annee <= annee_max)

    q = q.order_by(IdeCnuced.pays, IdeCnuced.annee)
    res = await db.execute(q)
    return [{"annee":r.annee, "valeur":float(r.valeur) if r.valeur is not None else None,
             "direction":r.direction, "indicateur":r.indicateur, "pays":r.pays}
            for r in res.scalars().all()]


# ── GET /ide/cnuced/pays-disponibles ─────────────────────────────────────────
@router.get("/cnuced/pays-disponibles")
async def get_pays_disponibles(db: AsyncSession = Depends(get_db)):
    """Retourne les pays qui ont des données IDE, avec leur code ISO"""
    from sqlalchemy import distinct
    from app.models.shared import RefPays
    res = await db.execute(select(distinct(IdeCnuced.pays)).order_by(IdeCnuced.pays))
    noms = [r[0] for r in res.fetchall()]
    # Enrichir avec les infos de ref_pays (code_iso2 pour le drapeau)
    result = []
    for nom in noms:
        rp = await db.execute(
            select(RefPays).where(
                (RefPays.nom_fr == nom) | (RefPays.nom_cnuced == nom)
            ).limit(1)
        )
        pays_obj = rp.scalar_one_or_none()
        result.append({
            "nom": nom,
            "code_iso2":   pays_obj.code_iso2   if pays_obj else None,
            "code_iso3":   pays_obj.code_iso3   if pays_obj else None,
            "continent":   pays_obj.continent   if pays_obj else None,
            "region_geo":  pays_obj.region_geo  if pays_obj else None,
        })
    return result


# ── GET /ide/cnuced/annees ────────────────────────────────────────────────────
# Bornes d'années réellement disponibles : la page publique s'aligne dessus
# automatiquement à chaque nouvel import (ex. millésime 2026).
@router.get("/cnuced/annees")
async def get_cnuced_annees(db: AsyncSession = Depends(get_db)):
    from sqlalchemy import func
    res = await db.execute(
        select(func.min(IdeCnuced.annee), func.max(IdeCnuced.annee))
        .where(IdeCnuced.valeur.isnot(None))
    )
    mn, mx = res.one()
    return {"annee_min": mn or 1990, "annee_max": mx or 2024}


# ── GET /ide/cnuced/kpis-calcules ─────────────────────────────────────────────
@router.get("/cnuced/kpis-calcules")
async def get_kpis_calcules(db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(IdeCnuced).where(IdeCnuced.pays=="Sénégal").order_by(IdeCnuced.annee))
    return calc_kpis(list(res.scalars().all()))


# ── GET /ide/kpis-config ──────────────────────────────────────────────────────
@router.get("/kpis-config")
async def get_kpis_config(db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(IdeKpiConfig).order_by(IdeKpiConfig.ordre))
    return [{"id":k.id,"code":k.code,"label":k.label,"description":k.description,"est_actif":k.est_actif,"ordre":k.ordre}
            for k in res.scalars().all()]


# ── GET /ide/kpis-config/actifs ───────────────────────────────────────────────
@router.get("/kpis-config/actifs")
async def get_kpis_actifs(db: AsyncSession = Depends(get_db)):
    res_conf = await db.execute(select(IdeKpiConfig).where(IdeKpiConfig.est_actif==True).order_by(IdeKpiConfig.ordre))
    configs  = res_conf.scalars().all()
    if not configs: return []
    if len(configs) > 4: configs = configs[:4]
    res_data = await db.execute(select(IdeCnuced).where(IdeCnuced.pays=="Sénégal").order_by(IdeCnuced.annee))
    kpis     = calc_kpis(list(res_data.scalars().all()))
    return [{"code":k.code,"label":k.label,"ordre":k.ordre,**kpis.get(k.code,{})} for k in configs]


# ── PATCH /ide/kpis-config/:id ────────────────────────────────────────────────
@router.patch("/kpis-config/{kpi_id}")
async def modifier_kpi_config(kpi_id: int, payload: dict, db: AsyncSession = Depends(get_db), current_user: dict = Depends(require_admin)):
    from fastapi import HTTPException
    res = await db.execute(select(IdeKpiConfig).where(IdeKpiConfig.id==kpi_id))
    k   = res.scalar_one_or_none()
    if not k: raise HTTPException(404, "KPI introuvable")
    if "est_actif" in payload: k.est_actif = payload["est_actif"]
    if "ordre"     in payload: k.ordre     = int(payload["ordre"])
    await db.flush()
    return {"id":k.id,"code":k.code,"label":k.label,"est_actif":k.est_actif,"ordre":k.ordre}


# ── GET /ide/analyses ─────────────────────────────────────────────────────────
@router.get("/analyses")
async def liste_analyses(
    source: Optional[str]  = Query(None),
    publie: Optional[bool] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    q = select(IdeAnalyse)
    if source: q = q.where(IdeAnalyse.source == source)
    if publie is not None: q = q.where(IdeAnalyse.est_publie == publie)
    q = q.order_by(IdeAnalyse.ordre, IdeAnalyse.created_at.desc())
    res = await db.execute(q)
    return [analyse_to_dict(a) for a in res.scalars().all()]


def analyse_to_dict(a: IdeAnalyse) -> dict:
    return {"id":a.id,"source":a.source,"titre":a.titre,"commentaire":a.commentaire,
            "direction":a.direction,"indicateur":a.indicateur,"annee_debut":a.annee_debut,
            "annee_fin":a.annee_fin,"est_publie":a.est_publie,"ordre":a.ordre,
            "created_at":a.created_at.isoformat() if a.created_at else None}


# ── POST /ide/analyses ────────────────────────────────────────────────────────
@router.post("/analyses", status_code=201)
async def creer_analyse(payload: dict, db: AsyncSession = Depends(get_db), current_user: dict = Depends(require_admin)):
    from fastapi import HTTPException
    if not payload.get("titre") or not payload.get("commentaire"):
        raise HTTPException(422, "Titre et commentaire obligatoires")
    a = IdeAnalyse(
        source      = payload.get("source","cnuced"),
        titre       = payload["titre"],
        commentaire = payload["commentaire"],
        direction   = payload.get("direction") or None,
        indicateur  = payload.get("indicateur") or None,
        annee_debut = int(payload["annee_debut"]) if payload.get("annee_debut") else None,
        annee_fin   = int(payload["annee_fin"])   if payload.get("annee_fin")   else None,
        ordre       = int(payload.get("ordre",0)),
    )
    db.add(a); await db.flush()
    return analyse_to_dict(a)


# ── PATCH /ide/analyses/:id ───────────────────────────────────────────────────
@router.patch("/analyses/{analyse_id}")
async def modifier_analyse(analyse_id: int, payload: dict, db: AsyncSession = Depends(get_db), current_user: dict = Depends(require_admin)):
    from fastapi import HTTPException
    res = await db.execute(select(IdeAnalyse).where(IdeAnalyse.id==analyse_id))
    a   = res.scalar_one_or_none()
    if not a: raise HTTPException(404, "Analyse introuvable")
    for f in ["titre","commentaire","direction","indicateur","source"]:
        if f in payload: setattr(a, f, payload[f] or None)
    for f in ["annee_debut","annee_fin"]:
        if f in payload: setattr(a, f, int(payload[f]) if payload[f] else None)
    if "ordre"     in payload: a.ordre     = int(payload.get("ordre",0))
    if "est_publie" in payload: a.est_publie = payload["est_publie"]
    await db.flush()
    return analyse_to_dict(a)


# ── DELETE /ide/analyses/:id ──────────────────────────────────────────────────
@router.delete("/analyses/{analyse_id}", status_code=204)
async def supprimer_analyse(analyse_id: int, db: AsyncSession = Depends(get_db), current_user: dict = Depends(require_admin)):
    from fastapi import HTTPException
    res = await db.execute(select(IdeAnalyse).where(IdeAnalyse.id==analyse_id))
    a   = res.scalar_one_or_none()
    if not a: raise HTTPException(404, "Analyse introuvable")
    await db.delete(a); await db.flush()


# ── GET /ide/pays-ref ─────────────────────────────────────────────────────────
@router.get("/pays-ref")
async def get_pays_ref(db: AsyncSession = Depends(get_db)):
    from app.models.shared import RefPays
    res = await db.execute(
        select(RefPays).where(RefPays.actif == True).order_by(RefPays.nom_fr)
    )
    return [
        {
            "id": p.id,
            "nom_fr": p.nom_fr,
            "nom_cnuced": p.nom_cnuced,
            "code_iso2": p.code_iso2,
            "continent": p.continent,
            "region_geo": p.region_geo,
            "niveau_revenu": p.niveau_revenu,
        }
        for p in res.scalars().all()
    ]


# ── Helpers import / parse ────────────────────────────────────────────────────
# Format CSV CNUCED : Economy_Label | Year | <serie>_Value  (1 ligne par année)

from fastapi import UploadFile, File, Form
from typing import List
import csv, io

# ── Helpers import / parse ────────────────────────────────────────────────────
# Format CSV CNUCED : Economy_Label | Year | <serie>_Value  (1 ligne par année)
# Supporte les fichiers mono-pays ET multi-pays dans le même fichier.

from fastapi import UploadFile, File, Form
from typing import List
import csv, io

def _parse_cnuced_file(contenu: bytes, nom_fichier: str) -> dict[str, list[tuple[int, float | None]]]:
    """Retourne {economy_label: [(annee, valeur), ...]} — supporte fichiers multi-pays."""
    ext = (nom_fichier or "").lower().rsplit(".", 1)[-1]

    if ext in ("xlsx", "xls"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(contenu), data_only=True)
        ws = wb.active
        data = list(ws.iter_rows(values_only=True))
    else:
        texte = contenu.decode("utf-8-sig", errors="replace")
        sep = "," if texte[:2048].count(",") >= texte[:2048].count(";") else ";"
        data = list(csv.reader(io.StringIO(texte), delimiter=sep))

    header_labels = {c.strip().lower() for c in (data[0] if data else []) if c}
    SKIP = {"economy_label", "economy", "économie", "pays"}

    result: dict[str, list[tuple[int, float | None]]] = {}
    for row in data[1:]:
        try:
            label = str(row[0]).strip() if row[0] else ""
            if not label or label.lower() in SKIP:
                continue
            annee = int(str(row[1]).strip())
            if not (1970 <= annee <= 2050):
                continue
            raw = str(row[2]).strip() if len(row) > 2 else ""
            valeur = None if raw in ("", "...", "—", "-", "n.d.", "N/A") else float(
                raw.replace(",", ".").replace(" ", "").replace("\xa0", "")
            )
            result.setdefault(label, []).append((annee, valeur))
        except (ValueError, IndexError):
            continue
    return result


# ── Parseur « Annex tables » (World Investment Report UNCTAD) ────────────────
# Format large : lignes de titre, puis en-tête `Region/economy | 1990 | … | 2025`,
# une ligne par économie avec les années en colonnes (millions USD). Les lignes
# d'agrégats régionaux et les notes de bas de page sont ignorées.

_ANNEX_AGREGATS = {
    "world", "developed economies", "developing economies", "memorandum",
    "europe", "european union", "other europe", "other developed europe",
    "north america", "other developed economies",
    "africa", "north africa", "other africa", "central africa", "east africa",
    "southern africa", "west africa",
    "asia", "east and south-east asia", "east asia", "south-east asia",
    "south asia", "west asia", "central asia",
    "latin america and the caribbean", "south and central america",
    "south america", "central america", "caribbean", "oceania",
    "least developed countries (ldcs)", "landlocked countries (llcs)",
    "landlocked developing countries (lldcs)",
    "small island developing states (sids)",
}


def _parse_annex_file(contenu: bytes, nom_fichier: str) -> dict[str, list[tuple[int, float | None]]]:
    """Retourne {economy_label: [(annee, valeur), ...]} depuis une Annex table UNCTAD."""
    ext = (nom_fichier or "").lower().rsplit(".", 1)[-1]

    if ext in ("xlsx", "xls"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(contenu), data_only=True)
        ws = wb.active
        data = list(ws.iter_rows(values_only=True))
    else:
        texte = contenu.decode("utf-8-sig", errors="replace")
        sep = "," if texte[:2048].count(",") >= texte[:2048].count(";") else ";"
        data = list(csv.reader(io.StringIO(texte), delimiter=sep))

    def _annee(cell) -> int | None:
        try:
            a = int(str(cell).strip())
            return a if 1900 <= a <= 2100 else None
        except (ValueError, TypeError):
            return None

    # En-tête : première ligne (parmi les 10 premières) avec au moins 5 années
    annees_cols: dict[int, int] = {}
    debut = 0
    for i, row in enumerate(data[:10]):
        cols = {ci: a for ci, cell in enumerate(row[1:], start=1) if (a := _annee(cell)) is not None}
        if len(cols) >= 5:
            annees_cols, debut = cols, i + 1
            break
    if not annees_cols:
        raise ValueError("en-tête introuvable (ligne « Region/economy | 1990 | … » attendue)")

    def _valeur(cell) -> float | None:
        if cell is None:
            return None
        if isinstance(cell, (int, float)):
            return float(cell)
        raw = str(cell).strip()
        if raw in ("", "...", "..", "—", "-", "n.d.", "N/A"):
            return None
        try:
            return float(raw.replace(",", ".").replace(" ", "").replace("\xa0", ""))
        except ValueError:
            return None

    result: dict[str, list[tuple[int, float | None]]] = {}
    for row in data[debut:]:
        label = str(row[0]).strip() if row and row[0] is not None else ""
        low = label.lower()
        if not label or low in _ANNEX_AGREGATS or low.startswith(("source:", "note:")):
            continue
        lignes = [(annee, _valeur(row[ci]) if ci < len(row) else None) for ci, annee in sorted(annees_cols.items(), key=lambda kv: kv[1])]
        if all(v is None for _, v in lignes):
            continue  # ligne vide / décorative
        result.setdefault(label, []).extend(lignes)
    return result


def _normalize_name(s: str) -> str:
    """Normalise un nom de pays : supprime accents, minuscules, apostrophes,
    et suffixes temporels UNCTAD comme ‘(...2011)’, ‘(..1991)’, ‘(...)’."""
    import unicodedata, re
    s = re.sub(r"\s*\(\.+\d*\)", "", s)   # "Ethiopie (...1991)" → "Ethiopie"
    s = s.replace("’", "’").replace("’", "’").replace("`", "’")
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.lower().strip()


async def _resolve_ref_pays(db, label: str):
    """Résout un nom UNCTAD → RefPays. Essaie d'abord exact, puis normalisé (sans accents)."""
    from app.models.shared import RefPays
    from sqlalchemy import or_

    # 1. Correspondance exacte
    res = await db.execute(
        select(RefPays).where(or_(RefPays.nom_cnuced == label, RefPays.nom_fr == label)).limit(1)
    )
    obj = res.scalar_one_or_none()
    if obj:
        return obj

    # 2. Correspondance normalisée (gère accents, apostrophes, casse)
    label_norm = _normalize_name(label)
    all_res = await db.execute(select(RefPays))
    for p in all_res.scalars().all():
        if _normalize_name(p.nom_fr or "") == label_norm or _normalize_name(p.nom_cnuced or "") == label_norm:
            return p

    return None


async def _upsert_serie(db, ref_pays_id: int, nom_pays: str, direction: str, indicateur: str,
                        lignes: list[tuple[int, float | None]]) -> tuple[int, int]:
    insere = mis_a_jour = 0
    for annee, valeur in lignes:
        res = await db.execute(
            select(IdeCnuced).where(
                IdeCnuced.ref_pays_id == ref_pays_id,
                IdeCnuced.annee == annee,
                IdeCnuced.direction == direction,
                IdeCnuced.indicateur == indicateur,
            )
        )
        row = res.scalar_one_or_none()
        if row:
            row.valeur = valeur
            row.pays = nom_pays
            mis_a_jour += 1
        else:
            db.add(IdeCnuced(
                pays=nom_pays, annee=annee, direction=direction, indicateur=indicateur,
                valeur=valeur, source="CNUCED", ref_pays_id=ref_pays_id,
            ))
            insere += 1
    return insere, mis_a_jour


async def _recalc_monde(db) -> None:
    """Recalcule les agrégats ide_cnuced_monde après une modification des
    données IDE : un UPDATE no-op sur ref_groupements déclenche le trigger
    sync_icm_on_groupement (migration 066), qui reconstruit les statistiques
    de chaque groupement — même logique que lors d'un changement de groupement,
    sans duplication de code."""
    from sqlalchemy import text
    await db.execute(text("UPDATE ref_groupements SET code = code"))


# ── POST /ide/importer ────────────────────────────────────────────────────────
# Chaque zone accepte N fichiers (un par pays). Le pays est auto-détecté
# depuis la colonne A (Economy_Label) et mappé vers ref_pays.

@router.post("/importer", status_code=200)
async def importer_ide(
    flux_entrant:  List[UploadFile] = File(default=[]),
    flux_sortant:  List[UploadFile] = File(default=[]),
    stock_entrant: List[UploadFile] = File(default=[]),
    stock_sortant: List[UploadFile] = File(default=[]),
    # Mode d'extraction : "annex" = Annex tables WIR (années en colonnes, défaut),
    # "series" = séries CNUCED historiques (Economy_Label | Year | Value).
    format_import: str = Form(default="series"),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_admin),
):
    from fastapi import HTTPException

    zones = {
        ("entrant", "flux"):  flux_entrant,
        ("sortant", "flux"):  flux_sortant,
        ("entrant", "stock"): stock_entrant,
        ("sortant", "stock"): stock_sortant,
    }

    resultats:   dict[str, dict] = {}   # nom_pays → stats
    erreurs:     list[str]       = []   # erreurs techniques (lecture fichier)
    non_resolus: dict[str, int]  = {}   # label UNCTAD → nb_lignes non importées

    for (direction, indicateur), fichiers in zones.items():
        for fichier in (fichiers or []):
            contenu = await fichier.read()
            if not contenu:
                continue

            parseur = _parse_annex_file if format_import == "annex" else _parse_cnuced_file
            try:
                lignes_par_pays = parseur(contenu, fichier.filename or "")
            except Exception as e:
                erreurs.append(f"{fichier.filename}: erreur de lecture — {e}")
                continue

            if not lignes_par_pays:
                erreurs.append(f"{fichier.filename}: aucune ligne valide trouvée")
                continue

            for label, lignes in lignes_par_pays.items():
                pays_obj = await _resolve_ref_pays(db, label)
                if not pays_obj:
                    non_resolus[label] = non_resolus.get(label, 0) + len(lignes)
                    continue

                ins, maj = await _upsert_serie(db, pays_obj.id, pays_obj.nom_fr, direction, indicateur, lignes)
                nom = pays_obj.nom_fr
                if nom not in resultats:
                    resultats[nom] = {"ref_pays_id": pays_obj.id, "insere": 0, "mis_a_jour": 0}
                resultats[nom]["insere"] += ins
                resultats[nom]["mis_a_jour"] += maj

    if resultats:
        await _recalc_monde(db)  # tenir la vue Monde à jour automatiquement
    await db.flush()
    return {
        "pays": [
            {"pays": nom, "ref_pays_id": d["ref_pays_id"], "insere": d["insere"], "mis_a_jour": d["mis_a_jour"]}
            for nom, d in sorted(resultats.items())
        ],
        "erreurs": erreurs,
        "non_resolus": [
            {"label": label, "nb_lignes": nb}
            for label, nb in sorted(non_resolus.items())
        ],
    }


# ── POST /ide/associer-pays ──────────────────────────────────────────────────
# Enregistre un alias UNCTAD → ref_pays en mettant à jour nom_cnuced.
# Permet de résoudre manuellement les pays non reconnus lors d'un import.

@router.post("/associer-pays", status_code=200)
async def associer_pays(payload: dict, db: AsyncSession = Depends(get_db), current_user: dict = Depends(require_admin)):
    from fastapi import HTTPException
    from app.models.shared import RefPays

    label   = (payload.get("label_cnuced") or "").strip()
    ref_id  = payload.get("ref_pays_id")

    if not label or not ref_id:
        raise HTTPException(400, "label_cnuced et ref_pays_id sont requis")

    res = await db.execute(select(RefPays).where(RefPays.id == int(ref_id)))
    p   = res.scalar_one_or_none()
    if not p:
        raise HTTPException(404, "Pays introuvable dans ref_pays")

    p.nom_cnuced = label
    await db.flush()
    return {"success": True, "nom_fr": p.nom_fr, "nom_cnuced": p.nom_cnuced}


# ── POST /ide/rafraichir ──────────────────────────────────────────────────────
# Fetch UNCTAD API pour tous les pays déjà importés. Nécessite les vars d'env
# UNCTAD_CLIENT_ID et UNCTAD_CLIENT_SECRET.
#
# API UNCTAD (OAuth2 client_credentials) :
#   POST https://unctadstat-user-api.unctad.org/token
#   GET  https://unctadstat-user-api.unctad.org/US.FdiFlowsStock/cur/Facts
#        ?$format=csv&culture=fr&$filter=Economy eq '{iso3}' and FlowType eq '{code}'

UNCTAD_TOKEN_URL = "https://unctadstat-user-api.unctad.org/token"
UNCTAD_DATA_URL  = "https://unctadstat-user-api.unctad.org/US.FdiFlowsStock/cur/Facts"

# Mapping code UNCTAD FlowType → (direction, indicateur)
UNCTAD_SERIES = [
    ("FDI_F_In",  "entrant", "flux"),
    ("FDI_F_Out", "sortant", "flux"),
    ("FDI_S_In",  "entrant", "stock"),
    ("FDI_S_Out", "sortant", "stock"),
]


async def _get_unctad_token(client_id: str, client_secret: str) -> str:
    import httpx
    async with httpx.AsyncClient(timeout=30) as client:
        r = await client.post(
            UNCTAD_TOKEN_URL,
            data={"grant_type": "client_credentials", "client_id": client_id, "client_secret": client_secret},
        )
        if not r.is_success:
            raise ValueError(f"HTTP {r.status_code} — réponse: {r.text[:500]}")
        return r.json()["access_token"]


async def _fetch_unctad_series(token: str, iso3: str, flow_type: str) -> bytes:
    import httpx
    async with httpx.AsyncClient(timeout=60) as client:
        r = await client.get(
            UNCTAD_DATA_URL,
            params={
                "$format": "csv",
                "culture":  "fr",
                "$filter":  f"Economy eq '{iso3}' and FlowType eq '{flow_type}'",
            },
            headers={"Authorization": f"Bearer {token}"},
        )
        r.raise_for_status()
        return r.content


async def _do_rafraichir(db_factory) -> dict:
    """Logique de refresh UNCTAD — appelée depuis la route ET le scheduler."""
    from app.core.database import AsyncSessionLocal
    from app.models.shared import RefPays
    from app.core.config import get_settings
    s = get_settings()
    client_id     = s.UNCTAD_CLIENT_ID
    client_secret = s.UNCTAD_CLIENT_SECRET
    if not client_id or not client_secret:
        return {"success": False, "erreur": "Variables UNCTAD_CLIENT_ID / UNCTAD_CLIENT_SECRET non configurées"}

    async with db_factory() as db:
        # Récupérer tous les pays déjà importés
        from sqlalchemy import distinct
        res = await db.execute(
            select(distinct(IdeCnuced.ref_pays_id)).where(IdeCnuced.ref_pays_id != None)
        )
        ref_pays_ids = [r[0] for r in res.fetchall()]

        if not ref_pays_ids:
            return {"success": True, "message": "Aucun pays importé", "pays": []}

        try:
            token = await _get_unctad_token(client_id, client_secret)
        except Exception as e:
            return {"success": False, "erreur": f"Authentification UNCTAD échouée : {e}"}

        resultats = []
        erreurs   = []

        for rpid in ref_pays_ids:
            rp = await db.execute(select(RefPays).where(RefPays.id == rpid))
            pays_obj = rp.scalar_one_or_none()
            if not pays_obj or not pays_obj.code_iso3:
                continue

            iso3    = pays_obj.code_iso3
            nom_fr  = pays_obj.nom_fr
            p_ins   = p_maj = 0

            for flow_type, direction, indicateur in UNCTAD_SERIES:
                try:
                    contenu     = await _fetch_unctad_series(token, iso3, flow_type)
                    lignes_dict = _parse_cnuced_file(contenu, f"{iso3}_{flow_type}.csv")
                    # Aplatir — on connaît déjà le pays pour le refresh
                    lignes = [item for items in lignes_dict.values() for item in items]
                    ins, maj = await _upsert_serie(db, rpid, nom_fr, direction, indicateur, lignes)
                    p_ins += ins
                    p_maj += maj
                except Exception as e:
                    erreurs.append(f"{nom_fr}/{flow_type}: {e}")

            resultats.append({"pays": nom_fr, "insere": p_ins, "mis_a_jour": p_maj})

        if resultats:
            await _recalc_monde(db)  # tenir la vue Monde à jour automatiquement
        await db.commit()
        return {"success": True, "pays": resultats, "erreurs": erreurs}


@router.post("/rafraichir", status_code=200)
async def rafraichir_ide(db: AsyncSession = Depends(get_db), current_user: dict = Depends(require_admin)):
    from app.core.database import AsyncSessionLocal
    return await _do_rafraichir(AsyncSessionLocal)


@router.get("/rafraichir/config")
async def rafraichir_config():
    """Indique si les credentials UNCTAD sont configurés."""
    from app.core.config import get_settings
    s = get_settings()
    ok = bool(s.UNCTAD_CLIENT_ID and s.UNCTAD_CLIENT_SECRET)
    return {"configured": ok}


# ── GET /ide/cnuced/stats ─────────────────────────────────────────────────────
@router.get("/cnuced/stats")
async def get_cnuced_stats(db: AsyncSession = Depends(get_db)):
    from sqlalchemy import func as sqlfunc, text
    from app.models.shared import RefPays

    res = await db.execute(
        select(
            IdeCnuced.ref_pays_id,
            IdeCnuced.pays,
            IdeCnuced.direction,
            IdeCnuced.indicateur,
            sqlfunc.min(IdeCnuced.annee).label("annee_min"),
            sqlfunc.max(IdeCnuced.annee).label("annee_max"),
            sqlfunc.count(IdeCnuced.id).label("nb"),
        )
        .where(IdeCnuced.ref_pays_id != None)
        .group_by(IdeCnuced.ref_pays_id, IdeCnuced.pays, IdeCnuced.direction, IdeCnuced.indicateur)
        .order_by(IdeCnuced.pays)
    )
    rows = res.fetchall()

    # Regrouper par pays
    pays_dict: dict = {}
    for row in rows:
        rpid = row.ref_pays_id
        if rpid not in pays_dict:
            pays_dict[rpid] = {
                "ref_pays_id": rpid,
                "pays": row.pays,
                "series": {},
            }
        key = f"{row.direction}_{row.indicateur}"
        pays_dict[rpid]["series"][key] = {
            "annee_min": row.annee_min,
            "annee_max": row.annee_max,
            "nb": row.nb,
        }

    # Enrichir avec code_iso2
    result = []
    for rpid, data in pays_dict.items():
        rp = await db.execute(select(RefPays).where(RefPays.id == rpid))
        pays_obj = rp.scalar_one_or_none()
        result.append({
            "ref_pays_id": rpid,
            "pays": data["pays"],
            "code_iso2": pays_obj.code_iso2 if pays_obj else None,
            "series": data["series"],
        })

    result.sort(key=lambda x: x["pays"])
    return result


# ── DELETE /ide/cnuced/pays/:ref_pays_id ─────────────────────────────────────
@router.delete("/cnuced/pays/{ref_pays_id}", status_code=200)
async def supprimer_pays_ide(ref_pays_id: int, db: AsyncSession = Depends(get_db), current_user: dict = Depends(require_admin)):
    from fastapi import HTTPException
    from sqlalchemy import delete as sqldel
    res = await db.execute(
        select(IdeCnuced).where(IdeCnuced.ref_pays_id == ref_pays_id).limit(1)
    )
    if not res.scalar_one_or_none():
        raise HTTPException(404, "Aucune donnée IDE pour ce pays")
    result = await db.execute(
        sqldel(IdeCnuced).where(IdeCnuced.ref_pays_id == ref_pays_id)
    )
    await _recalc_monde(db)  # tenir la vue Monde à jour automatiquement
    await db.flush()
    return {"success": True, "supprime": result.rowcount}
