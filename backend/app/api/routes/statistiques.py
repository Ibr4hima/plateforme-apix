"""
Module Statistiques : données macro par pays (population, superficie, PIB…),
séries temporelles, analyse par pays / comparative, et fiches de comparaison.
Les indicateurs « dérivés » (densité, PIB/hab) sont calculés à la volée pour
rester toujours cohérents avec les valeurs de base.
"""
from typing import List, Optional
from fastapi import APIRouter, Depends, Query
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.models.shared import RefPays, StatIndicateur, StatPays
from app.core.uploads import lire_import

router = APIRouter(prefix="/statistiques", tags=["Statistiques"])


@router.get("/indicateurs")
async def indicateurs(db: AsyncSession = Depends(get_db)):
    rows = (await db.execute(select(StatIndicateur).order_by(StatIndicateur.ordre))).scalars().all()
    return [{"code": r.code, "libelle": r.libelle, "unite": r.unite,
             "categorie": r.categorie, "ordre": r.ordre, "derive": r.derive} for r in rows]


@router.get("/pays")
async def pays_disponibles(db: AsyncSession = Depends(get_db)):
    """Pays ayant au moins une donnée statistique."""
    ids = (await db.execute(select(StatPays.pays_id).distinct())).scalars().all()
    if not ids:
        return []
    rows = (await db.execute(
        select(RefPays).where(RefPays.id.in_(ids)).order_by(RefPays.nom_fr)
    )).scalars().all()
    return [{"id": p.id, "nom": p.nom_fr, "code_iso3": p.code_iso3, "code_iso2": p.code_iso2,
             "continent": p.continent, "region_geo": p.region_geo} for p in rows]


def _completer_derives(par_pays_annee: dict) -> None:
    """Ajoute densité (pop/superficie) et pib_hab (pib*1e9/pop) là où c'est possible."""
    for (pid, annee), vals in par_pays_annee.items():
        pop = vals.get("population")
        surf = vals.get("superficie")
        pib = vals.get("pib")
        if pop and surf and surf > 0:
            vals["densite"] = round(pop / surf, 2)
        if pop and pib and pop > 0:
            vals["pib_hab"] = round(pib / pop, 1)  # pib en USD bruts
        im, ex = vals.get("importations_marchandises"), vals.get("exportations_marchandises")
        if im is not None and ex is not None:
            vals["balance_marchandises"] = round(ex - im, 1)
        ims, exs = vals.get("importations_services"), vals.get("exportations_services")
        if ims is not None and exs is not None:
            vals["balance_services"] = round(exs - ims, 1)


async def _donnees(db: AsyncSession, pays_ids: List[int],
                   annee_min: Optional[int], annee_max: Optional[int]) -> List[dict]:
    q = select(StatPays).where(StatPays.pays_id.in_(pays_ids))
    if annee_min is not None:
        q = q.where(StatPays.annee >= annee_min)
    if annee_max is not None:
        q = q.where(StatPays.annee <= annee_max)
    rows = (await db.execute(q)).scalars().all()
    # noms des pays
    noms = {p.id: p.nom_fr for p in (await db.execute(
        select(RefPays).where(RefPays.id.in_(pays_ids)))).scalars().all()}
    # Superficie : constante par pays (stockée à l'année sentinelle) → on la
    # propage à toutes les années réelles du pays pour densité/KPI/graphes.
    superf = {}
    for r in rows:
        if r.indicateur == "superficie" and r.valeur is not None:
            superf[r.pays_id] = float(r.valeur)  # une seule valeur par pays
    par = {}
    for r in rows:
        if r.indicateur == "superficie":
            continue
        par.setdefault((r.pays_id, r.annee), {})[r.indicateur] = float(r.valeur) if r.valeur is not None else None
    for (pid, annee), vals in par.items():
        if pid in superf:
            vals["superficie"] = superf[pid]
    _completer_derives(par)
    out = []
    for (pid, annee), vals in par.items():
        for ind, val in vals.items():
            out.append({"pays_id": pid, "pays": noms.get(pid, ""), "annee": annee,
                        "indicateur": ind, "valeur": val})
    return out


@router.get("/donnees")
async def donnees(
    pays: str = Query(..., description="ids de pays séparés par virgule"),
    annee_min: Optional[int] = None,
    annee_max: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
):
    ids = [int(x) for x in pays.split(",") if x.strip().isdigit()]
    if not ids:
        return []
    return await _donnees(db, ids, annee_min, annee_max)


@router.get("/comparaison")
async def comparaison(
    pays: str = Query(..., description="ids de pays à comparer, séparés par virgule"),
    annee: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
):
    """Fiche de comparaison : dernière valeur (ou année demandée) de chaque
    indicateur pour chaque pays, prête pour un tableau côte à côte."""
    ids = [int(x) for x in pays.split(",") if x.strip().isdigit()]
    if not ids:
        return {"pays": [], "indicateurs": [], "valeurs": {}}
    inds = (await db.execute(select(StatIndicateur).order_by(StatIndicateur.ordre))).scalars().all()
    data = await _donnees(db, ids, None, None)
    # dernière année disponible par (pays, indicateur), ou l'année demandée
    latest = {}
    for d in data:
        if annee is not None and d["annee"] != annee:
            continue
        key = (d["pays_id"], d["indicateur"])
        if key not in latest or d["annee"] > latest[key]["annee"]:
            latest[key] = d
    paysrows = (await db.execute(select(RefPays).where(RefPays.id.in_(ids)))).scalars().all()
    paysmap = {p.id: p for p in paysrows}
    valeurs = {}
    for (pid, ind), d in latest.items():
        valeurs.setdefault(str(pid), {})[ind] = {"valeur": d["valeur"], "annee": d["annee"]}
    return {
        "pays": [{"id": pid, "nom": paysmap[pid].nom_fr, "code_iso3": paysmap[pid].code_iso3,
                  "code_iso2": paysmap[pid].code_iso2}
                 for pid in ids if pid in paysmap],
        "indicateurs": [{"code": i.code, "libelle": i.libelle, "unite": i.unite,
                         "categorie": i.categorie, "ordre": i.ordre} for i in inds],
        "valeurs": valeurs,
    }


@router.get("/ide_flux")
async def ide_flux(
    pays: str = Query(..., description="ids de pays séparés par virgule"),
    indicateur: str = Query("flux", pattern="^(flux|stock)$"),
    db: AsyncSession = Depends(get_db),
):
    """Dernière valeur d'IDE entrante et sortante (source CNUCED) par pays,
    en USD — flux par défaut, stock via ?indicateur=stock. Les valeurs CNUCED
    sont en millions USD → converties en USD (×1e6)."""
    from app.models.ide import IdeCnuced
    ids = [int(x) for x in pays.split(",") if x.strip().isdigit()]
    if not ids:
        return {}
    rows = (await db.execute(
        select(IdeCnuced).where(IdeCnuced.ref_pays_id.in_(ids), IdeCnuced.indicateur == indicateur)
    )).scalars().all()
    latest: dict = {}
    for r in rows:
        if r.ref_pays_id is None or r.valeur is None:
            continue
        key = (r.ref_pays_id, r.direction)
        if key not in latest or r.annee > latest[key][1]:
            latest[key] = (float(r.valeur), r.annee)
    out: dict = {}
    for (pid, direction), (v, an) in latest.items():
        out.setdefault(str(pid), {})[direction] = {"valeur": v * 1_000_000, "annee": an}
    return out


# ══════════════════════════════════════════════════════════════════════════════
# IMPORT ADMIN — extraction des données depuis des fichiers Excel/CSV
# Format attendu : colonne A = pays, colonne B = année, colonne C = valeur.
# Un fichier peut contenir plusieurs pays. Le pays est auto-résolu vers ref_pays
# (exact puis normalisé sans accents), avec possibilité d'associer manuellement
# les libellés non reconnus (comme pour l'IDE).
# ══════════════════════════════════════════════════════════════════════════════
import io
import csv
import unicodedata
import re as _re
from fastapi import UploadFile, File, Form, HTTPException
from app.core.auth import require_admin
from app.models.shared import StatPays as _StatPays

# Seuls les indicateurs « de base » sont importables (densité et PIB/hab dérivés)
_INDICATEURS_IMPORT = {
    "population", "superficie", "pib", "croissance_pib",
    "importations_marchandises", "exportations_marchandises",
    "importations_services", "exportations_services",
}
# Indicateurs exprimés en millions de USD dans les fichiers → stockés en USD (×1e6)
_MILLIONS_USD = {
    "pib", "importations_marchandises", "exportations_marchandises",
    "importations_services", "exportations_services",
}


def _lire_tableur(contenu: bytes, nom_fichier: str):
    ext = (nom_fichier or "").lower().rsplit(".", 1)[-1]
    if ext in ("xlsx", "xls"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(contenu), data_only=True)
        return list(wb.active.iter_rows(values_only=True))
    texte = contenu.decode("utf-8-sig", errors="replace")
    sep = "," if texte[:2048].count(",") >= texte[:2048].count(";") else ";"
    return list(csv.reader(io.StringIO(texte), delimiter=sep))


def _detecter_colonnes(header):
    """Repère les colonnes Pays / Année / Valeur depuis l'en-tête (ordre libre).
    Repli positionnel : Pays=0, Année=1, Valeur=2 si l'en-tête n'est pas reconnu."""
    hs = [str(c or "").strip().lower() for c in header]

    def trouver(mots):
        for i, h in enumerate(hs):
            if any(m in h for m in mots):
                return i
        return None

    pays = trouver(["pays", "country", "economy", "économie", "economie", "nom"])
    annee = trouver(["année", "annee", "year"])
    id_col = trouver(["id"])
    val = None
    for i in range(len(hs)):
        if i not in (pays, annee, id_col):
            val = i
            break
    if pays is None or annee is None or val is None:
        return 0, 1, 2
    return pays, annee, val


def _num(raw: str):
    if raw in ("", "...", "—", "-", "n.d.", "N/A", "None"):
        return None
    return float(raw.replace(" ", "").replace("\xa0", "").replace(",", "."))


def _parse_stat_file(contenu: bytes, nom_fichier: str):
    """Retourne {pays_label: [(annee, valeur), ...]} — colonnes détectées par
    en-tête (Pays / Année / Valeur dans n'importe quel ordre), multi-pays."""
    data = _lire_tableur(contenu, nom_fichier)
    if not data:
        return {}
    ci_pays, ci_annee, ci_val = _detecter_colonnes(data[0])
    SKIP = {"economy_label", "economy", "économie", "pays", "country", "nom", "libelle"}
    result = {}
    for row in data[1:]:
        try:
            label = str(row[ci_pays]).strip() if len(row) > ci_pays and row[ci_pays] else ""
            if not label or label.lower() in SKIP:
                continue
            annee = int(float(str(row[ci_annee]).strip()))
            if not (1900 <= annee <= 2100):
                continue
            raw = str(row[ci_val]).strip() if len(row) > ci_val else ""
            valeur = _num(raw)
            result.setdefault(label, []).append((annee, valeur))
        except (ValueError, IndexError):
            continue
    return result


# La superficie n'a pas d'année (colonnes : ID, Pays, Superficie) → année sentinelle.
ANNEE_SUPERFICIE = 0


def _parse_superficie_file(contenu: bytes, nom_fichier: str):
    """Retourne {pays_label: [(0, superficie)]} depuis un fichier ID / Pays / Superficie."""
    ext = (nom_fichier or "").lower().rsplit(".", 1)[-1]
    if ext in ("xlsx", "xls"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(contenu), data_only=True)
        data = list(wb.active.iter_rows(values_only=True))
    else:
        texte = contenu.decode("utf-8-sig", errors="replace")
        sep = "," if texte[:2048].count(",") >= texte[:2048].count(";") else ";"
        data = list(csv.reader(io.StringIO(texte), delimiter=sep))
    SKIP = {"pays", "country", "nom", "libelle", "economy_label"}
    result = {}
    for row in data[1:] if data else []:
        try:
            label = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if not label or label.lower() in SKIP:
                continue
            raw = str(row[2]).strip() if len(row) > 2 else ""
            valeur = None if raw in ("", "...", "—", "-", "n.d.", "N/A", "None") else float(
                raw.replace(" ", "").replace(",", ".").replace(" ", "")
            )
            result.setdefault(label, []).append((ANNEE_SUPERFICIE, valeur))
        except (ValueError, IndexError):
            continue
    return result


def _norm(s: str) -> str:
    s = _re.sub(r"\s*\(\.+\d*\)", "", s or "")
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.lower().strip()


async def _resolve_pays(db: AsyncSession, label: str, tous: list | None = None):
    from sqlalchemy import or_
    res = await db.execute(select(RefPays).where(or_(RefPays.nom_cnuced == label, RefPays.nom_fr == label, RefPays.code_iso3 == label.upper())).limit(1))
    obj = res.scalar_one_or_none()
    if obj:
        return obj
    ln = _norm(label)
    if tous is None:
        tous = (await db.execute(select(RefPays))).scalars().all()
    for p in tous:
        if _norm(p.nom_fr or "") == ln or _norm(p.nom_cnuced or "") == ln:
            return p
    return None


@router.get("/pays-ref")
async def pays_ref(db: AsyncSession = Depends(get_db)):
    """Tous les pays du référentiel (pour l'association manuelle)."""
    rows = (await db.execute(select(RefPays).order_by(RefPays.nom_fr))).scalars().all()
    return [{"id": p.id, "nom_fr": p.nom_fr, "code_iso3": p.code_iso3} for p in rows]


@router.get("/admin/couverture")
async def couverture(db: AsyncSession = Depends(get_db), _: dict = Depends(require_admin)):
    """Pour chaque pays ayant des données : indicateurs et plages d'années."""
    rows = (await db.execute(select(_StatPays))).scalars().all()
    noms = {p.id: (p.nom_fr, p.code_iso3) for p in (await db.execute(select(RefPays))).scalars().all()}
    agg = {}
    for r in rows:
        e = agg.setdefault(r.pays_id, {})
        s = e.setdefault(r.indicateur, {"min": r.annee, "max": r.annee, "nb": 0})
        s["min"] = min(s["min"], r.annee); s["max"] = max(s["max"], r.annee); s["nb"] += 1
    return [
        {"pays_id": pid, "pays": noms.get(pid, ("", ""))[0], "code_iso3": noms.get(pid, ("", ""))[1], "series": e}
        for pid, e in sorted(agg.items(), key=lambda x: noms.get(x[0], ("",))[0])
    ]


@router.post("/importer")
async def importer(
    indicateur: str = Form(...),
    fichiers: List[UploadFile] = File(default=[]),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_admin),
):
    if indicateur not in _INDICATEURS_IMPORT:
        raise HTTPException(400, f"Indicateur « {indicateur} » non importable (densité et PIB/hab sont calculés automatiquement).")

    resultats, erreurs, non_resolus = {}, [], {}
    # Référentiel préchargé une fois pour la résolution normalisée des labels
    tous_pays = (await db.execute(select(RefPays))).scalars().all()
    for fichier in (fichiers or []):
        contenu = await lire_import(fichier)
        if not contenu:
            continue
        try:
            par_pays = (_parse_superficie_file if indicateur == "superficie" else _parse_stat_file)(contenu, fichier.filename or "")
        except Exception as e:
            erreurs.append(f"{fichier.filename}: erreur de lecture — {e}")
            continue
        if not par_pays:
            erreurs.append(f"{fichier.filename}: aucune ligne valide trouvée")
            continue
        for label, lignes in par_pays.items():
            pays_obj = await _resolve_pays(db, label, tous_pays)
            if not pays_obj:
                non_resolus[label] = non_resolus.get(label, 0) + len(lignes)
                continue
            # Conversion d'unités puis upsert en masse sur la PK (pays_id, annee, indicateur)
            par_annee: dict = {}
            for annee, valeur in lignes:
                if valeur is not None:
                    if indicateur == "population":
                        valeur = valeur * 1000            # milliers d'habitants → habitants
                    elif indicateur in _MILLIONS_USD:
                        valeur = valeur * 1_000_000        # millions USD → USD
                par_annee[annee] = valeur
            annees = list(par_annee.keys())
            existants = set((await db.execute(select(_StatPays.annee).where(
                _StatPays.pays_id == pays_obj.id, _StatPays.indicateur == indicateur, _StatPays.annee.in_(annees)
            ))).scalars().all()) if annees else set()
            if annees:
                from sqlalchemy.dialects.postgresql import insert as pg_insert
                stmt = pg_insert(_StatPays).values([
                    {"pays_id": pays_obj.id, "annee": annee, "indicateur": indicateur, "valeur": valeur}
                    for annee, valeur in par_annee.items()
                ])
                stmt = stmt.on_conflict_do_update(
                    index_elements=["pays_id", "annee", "indicateur"],
                    set_={"valeur": stmt.excluded.valeur},
                )
                await db.execute(stmt)
            ins = sum(1 for a in annees if a not in existants); maj = len(annees) - ins
            d = resultats.setdefault(pays_obj.nom_fr, {"pays_id": pays_obj.id, "insere": 0, "mis_a_jour": 0})
            d["insere"] += ins; d["mis_a_jour"] += maj
    await db.flush()
    return {
        "pays": [{"pays": nom, "pays_id": d["pays_id"], "insere": d["insere"], "mis_a_jour": d["mis_a_jour"]} for nom, d in sorted(resultats.items())],
        "erreurs": erreurs,
        "non_resolus": [{"label": l, "nb_lignes": n} for l, n in sorted(non_resolus.items())],
    }


@router.post("/associer-pays")
async def associer_pays(payload: dict, db: AsyncSession = Depends(get_db), current_user: dict = Depends(require_admin)):
    label = (payload.get("label") or "").strip()
    ref_id = payload.get("ref_pays_id")
    if not label or not ref_id:
        raise HTTPException(400, "label et ref_pays_id sont requis")
    p = (await db.execute(select(RefPays).where(RefPays.id == int(ref_id)))).scalar_one_or_none()
    if not p:
        raise HTTPException(404, "Pays introuvable dans le référentiel")
    p.nom_cnuced = label
    await db.flush()
    return {"success": True, "nom_fr": p.nom_fr}


@router.delete("/indicateur/{code}", status_code=204)
async def vider_indicateur(code: str, db: AsyncSession = Depends(get_db), current_user: dict = Depends(require_admin)):
    """Supprime toutes les données d'un indicateur, pour tous les pays."""
    for r in (await db.execute(select(_StatPays).where(_StatPays.indicateur == code))).scalars().all():
        await db.delete(r)
    await db.flush()


@router.delete("/pays/{pays_id}", status_code=204)
async def supprimer_pays(pays_id: int, db: AsyncSession = Depends(get_db), current_user: dict = Depends(require_admin)):
    for r in (await db.execute(select(_StatPays).where(_StatPays.pays_id == pays_id))).scalars().all():
        await db.delete(r)
    await db.flush()


# ══════════════════════════════════════════════════════════════════════════════
# DONNÉES TRANSACTIONNELLES (resourcetrade.earth) — commerce bilatéral par ressource
# Colonnes du fichier Excel : Exporter ISO3, Importer ISO3, Resource, Year,
# Value (1000USD). Pays résolus par code ISO3 (secours : nom). Valeur ×1000 (le
# fichier est en milliers de USD). Libellés de ressources éditables.
# ══════════════════════════════════════════════════════════════════════════════
from app.models.shared import StatRessource, StatTransaction
from sqlalchemy import delete as _delete, func as _sqlfunc
from sqlalchemy.dialects.postgresql import insert as _pg_insert


async def _cache_pays(db: AsyncSession):
    rows = (await db.execute(select(RefPays))).scalars().all()
    by_code, by_name = {}, {}
    for p in rows:
        if p.code_iso3:
            by_code[p.code_iso3.upper()] = p.id
        if p.nom_fr:
            by_name[_norm(p.nom_fr)] = p.id
        if p.nom_cnuced:
            by_name[_norm(p.nom_cnuced)] = p.id
    return by_code, by_name


def _colonnes_tx(header):
    """Repère les colonnes du fichier resourcetrade (en-têtes anglais)."""
    hs = [str(c or "").strip().lower() for c in header]

    def trouver(*mots, exact=None):
        if exact is not None:
            for i, h in enumerate(hs):
                if h == exact:
                    return i
        for i, h in enumerate(hs):
            if all(m in h for m in mots):
                return i
        return None

    return {
        "exp_code": trouver("exporter", "iso"),
        "imp_code": trouver("importer", "iso"),
        "exp_nom": trouver(exact="exporter"),
        "imp_nom": trouver(exact="importer"),
        "ressource": trouver("resource") if trouver("resource") is not None else trouver("produit"),
        "annee": trouver("year") if trouver("year") is not None else trouver("année"),
        "valeur": trouver("value"),
    }


@router.post("/transactions/importer")
async def importer_transactions(
    fichiers: List[UploadFile] = File(default=[]),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_admin),
):
    """Importe un fichier Excel/CSV de transactions bilatérales par ressource.
    Réimporter une année remplace ses données."""
    import openpyxl
    by_code, by_name = await _cache_pays(db)
    crees: dict[str, str] = {}   # nom → code, partenaires créés automatiquement
    ressources: set[str] = set()
    annees_vues: set[int] = set()
    total = 0
    BATCH = 4000

    async def _resoudre_ou_creer(code: str, nom: str):
        """Résout un acteur (code ISO3 puis nom) ; s'il est absent, le crée pour
        ne perdre aucune donnée. Retourne son id, ou None si ni code ni nom."""
        code = (code or "").strip().upper()
        nom = (nom or "").strip()
        if code and code in by_code:
            return by_code[code]
        n = _norm(nom)
        if n and n in by_name:
            return by_name[n]
        if not code and not nom:
            return None
        # Création : code ISO3 s'il est libre et alphabétique, sinon sans code
        code_ok = code if (2 <= len(code) <= 3 and code.isalpha() and code not in by_code) else None
        p = RefPays(code_iso3=code_ok, nom_fr=nom or code, actif=False, origine="transaction")
        db.add(p)
        await db.flush()
        if code_ok:
            by_code[code_ok] = p.id
        if n:
            by_name[n] = p.id
        crees[nom or code] = code_ok or ""
        return p.id

    for fichier in (fichiers or []):
        contenu = await lire_import(fichier)
        if not contenu:
            continue
        ext = (fichier.filename or "").lower().rsplit(".", 1)[-1]
        if ext in ("xlsx", "xls"):
            wb = openpyxl.load_workbook(io.BytesIO(contenu), read_only=True, data_only=True)
            it = wb.active.iter_rows(values_only=True)
        else:
            texte = contenu.decode("utf-8-sig", errors="replace")
            sep = "," if texte[:2048].count(",") >= texte[:2048].count(";") else ";"
            it = iter(csv.reader(io.StringIO(texte), delimiter=sep))

        try:
            header = next(it)
        except StopIteration:
            continue
        c = _colonnes_tx(header)
        if c["exp_code"] is None or c["imp_code"] is None or c["valeur"] is None or c["annee"] is None:
            raise HTTPException(400, "Colonnes attendues introuvables (Exporter ISO3, Importer ISO3, Resource, Year, Value).")

        annees_purgees: set[int] = set()
        batch: list[dict] = []

        def g(row, key):
            i = c[key]
            return row[i] if i is not None and i < len(row) else None

        for row in it:
            try:
                annee = int(float(str(g(row, "annee")).strip()))
                if not (1900 <= annee <= 2100):
                    continue
                if annee not in annees_purgees:
                    await db.execute(_delete(StatTransaction).where(StatTransaction.annee == annee))
                    annees_purgees.add(annee); annees_vues.add(annee)

                eid = await _resoudre_ou_creer(str(g(row, "exp_code") or ""), str(g(row, "exp_nom") or ""))
                iid = await _resoudre_ou_creer(str(g(row, "imp_code") or ""), str(g(row, "imp_nom") or ""))
                if eid is None or iid is None:
                    continue

                ress = str(g(row, "ressource") or "").strip()
                if ress:
                    ressources.add(ress)
                val = _num(str(g(row, "valeur") or ""))
                if val is not None:
                    val = val * 1000  # milliers USD → USD

                batch.append({"annee": annee, "exportateur_id": eid, "importateur_id": iid,
                              "ressource": ress or None, "valeur": val})
                total += 1
                if len(batch) >= BATCH:
                    await db.execute(_pg_insert(StatTransaction.__table__), batch); batch = []
            except (ValueError, TypeError):
                continue
        if batch:
            await db.execute(_pg_insert(StatTransaction.__table__), batch)

    if ressources:
        await db.execute(
            _pg_insert(StatRessource.__table__).values([{"nom_en": r, "libelle": r} for r in ressources])
            .on_conflict_do_nothing(index_elements=["nom_en"])
        )
    await db.flush()
    return {
        "lignes": total,
        "annees": sorted(annees_vues),
        "ressources_vues": len(ressources),
        "partenaires_crees": [{"nom": nom, "code": c} for nom, c in sorted(crees.items())],
    }


@router.get("/transactions/couverture")
async def couverture_transactions(db: AsyncSession = Depends(get_db), _: dict = Depends(require_admin)):
    rows = (await db.execute(
        select(StatTransaction.annee, _sqlfunc.count()).group_by(StatTransaction.annee).order_by(StatTransaction.annee)
    )).all()
    return [{"annee": a, "nb_lignes": n} for a, n in rows]


@router.get("/transactions")
async def liste_transactions(
    db: AsyncSession = Depends(get_db),
    _: dict = Depends(require_admin),
    annee: Optional[int] = Query(default=None),
    exportateur_id: Optional[int] = Query(default=None),
    importateur_id: Optional[int] = Query(default=None),
    ressource: Optional[str] = Query(default=None),
    recherche: Optional[str] = Query(default=None),
    page: int = Query(default=1, ge=1),
    taille: int = Query(default=50, ge=1, le=500),
):
    """Tableau paginé des transactions importées (exportateur, importateur, année, ressource, valeur)."""
    from sqlalchemy import or_ as _or
    from sqlalchemy.orm import aliased as _aliased

    Exp = _aliased(RefPays)
    Imp = _aliased(RefPays)

    filtres = []
    if annee is not None:
        filtres.append(StatTransaction.annee == annee)
    if exportateur_id is not None:
        filtres.append(StatTransaction.exportateur_id == exportateur_id)
    if importateur_id is not None:
        filtres.append(StatTransaction.importateur_id == importateur_id)
    if ressource:
        filtres.append(StatTransaction.ressource == ressource)

    base = (
        select(
            StatTransaction.id, StatTransaction.annee, StatTransaction.valeur,
            StatTransaction.ressource,
            Exp.nom_fr.label("exp_nom"), Imp.nom_fr.label("imp_nom"),
            StatTransaction.exportateur_id, StatTransaction.importateur_id,
            StatRessource.libelle.label("ressource_lib"),
        )
        .join(Exp, Exp.id == StatTransaction.exportateur_id)
        .join(Imp, Imp.id == StatTransaction.importateur_id)
        .outerjoin(StatRessource, StatRessource.nom_en == StatTransaction.ressource)
    )
    for f in filtres:
        base = base.where(f)
    if recherche:
        motif = f"%{recherche.strip()}%"
        base = base.where(_or(Exp.nom_fr.ilike(motif), Imp.nom_fr.ilike(motif),
                              StatTransaction.ressource.ilike(motif),
                              StatRessource.libelle.ilike(motif)))

    sous = base.subquery()
    total = (await db.execute(select(_sqlfunc.count()).select_from(sous))).scalar_one()

    rows = (await db.execute(
        base.order_by(StatTransaction.annee.desc(), _sqlfunc.coalesce(StatRessource.libelle, StatTransaction.ressource),
                      Exp.nom_fr, Imp.nom_fr)
            .offset((page - 1) * taille).limit(taille)
    )).all()

    return {
        "total": total,
        "page": page,
        "taille": taille,
        "lignes": [
            {
                "id": r.id, "annee": r.annee,
                "exportateur": r.exp_nom, "importateur": r.imp_nom,
                "exportateur_id": r.exportateur_id, "importateur_id": r.importateur_id,
                "ressource": r.ressource_lib or r.ressource,
                "valeur": float(r.valeur) if r.valeur is not None else None,
            }
            for r in rows
        ],
    }


@router.delete("/transactions/{annee}", status_code=204)
async def supprimer_annee_transactions(annee: int, db: AsyncSession = Depends(get_db), current_user: dict = Depends(require_admin)):
    await db.execute(_delete(StatTransaction).where(StatTransaction.annee == annee))
    await db.flush()


@router.get("/partenaires")
async def liste_partenaires(db: AsyncSession = Depends(get_db), _: dict = Depends(require_admin)):
    """Partenaires commerciaux créés à l'import (territoires/agrégats hors référentiel)."""
    rows = (await db.execute(select(RefPays).where(RefPays.origine == "transaction").order_by(RefPays.nom_fr))).scalars().all()
    return [{"id": p.id, "nom_fr": p.nom_fr, "code_iso3": p.code_iso3} for p in rows]


@router.patch("/partenaires/{pays_id}")
async def modifier_partenaire(pays_id: int, payload: dict, db: AsyncSession = Depends(get_db), current_user: dict = Depends(require_admin)):
    p = (await db.execute(select(RefPays).where(RefPays.id == pays_id))).scalar_one_or_none()
    if not p or p.origine != "transaction":
        raise HTTPException(404, "Partenaire introuvable")
    if "nom_fr" in payload:
        nouveau = (payload["nom_fr"] or "").strip()
        if nouveau:
            # Conserve le nom d'origine comme alias pour que les imports futurs
            # (fichier en anglais) continuent de reconnaître ce partenaire.
            if not p.nom_cnuced:
                p.nom_cnuced = p.nom_fr
            p.nom_fr = nouveau
    await db.flush()
    return {"id": p.id, "nom_fr": p.nom_fr, "code_iso3": p.code_iso3}


@router.get("/ressources")
async def liste_ressources(db: AsyncSession = Depends(get_db), _: dict = Depends(require_admin)):
    rows = (await db.execute(select(StatRessource).order_by(StatRessource.nom_en))).scalars().all()
    return [{"nom_en": r.nom_en, "libelle": r.libelle} for r in rows]


@router.patch("/ressources/{nom_en:path}")
async def modifier_ressource(nom_en: str, payload: dict, db: AsyncSession = Depends(get_db), current_user: dict = Depends(require_admin)):
    r = (await db.execute(select(StatRessource).where(StatRessource.nom_en == nom_en))).scalar_one_or_none()
    if not r:
        raise HTTPException(404, "Ressource introuvable")
    if "libelle" in payload:
        r.libelle = (payload["libelle"] or "").strip() or r.nom_en
    await db.flush()
    return {"nom_en": r.nom_en, "libelle": r.libelle}


# ── Données commerciales — endpoints publics (page Statistiques) ───────────────

@router.get("/commerce/filtres")
async def commerce_filtres(db: AsyncSession = Depends(get_db)):
    """Options de filtre pour la vue publique : années, ressources, partenaires."""
    annees = (await db.execute(
        select(StatTransaction.annee).distinct().order_by(StatTransaction.annee.desc())
    )).scalars().all()
    ressources = (await db.execute(select(StatRessource).order_by(StatRessource.nom_en))).scalars().all()
    # Pays impliqués dans au moins une transaction (exportateur ou importateur)
    ids_exp = select(StatTransaction.exportateur_id.label("pid"))
    ids_imp = select(StatTransaction.importateur_id.label("pid"))
    ids = (await db.execute(select(ids_exp.union(ids_imp).subquery().c.pid))).scalars().all()
    pays = []
    if ids:
        rows = (await db.execute(select(RefPays).where(RefPays.id.in_(ids)).order_by(RefPays.nom_fr))).scalars().all()
        pays = [{"id": p.id, "nom": p.nom_fr, "code_iso3": p.code_iso3,
                 "continent": p.continent, "region_geo": p.region_geo} for p in rows]
    return {
        "annees": list(annees),
        "ressources": [{"nom_en": r.nom_en, "libelle": r.libelle or r.nom_en} for r in ressources],
        "pays": pays,
    }


@router.get("/commerce/detail")
async def commerce_detail(
    db: AsyncSession = Depends(get_db),
    pays_id: int = Query(...),
    direction: str = Query(default="exportateur"),
    annee: int = Query(...),
):
    """Détail groupé pour une année : par partenaire puis par ressource."""
    from sqlalchemy import and_ as _and
    from collections import defaultdict

    self_col = StatTransaction.exportateur_id if direction == "exportateur" else StatTransaction.importateur_id
    partner_col = StatTransaction.importateur_id if direction == "exportateur" else StatTransaction.exportateur_id
    W = _and(self_col == pays_id, StatTransaction.annee == annee)
    _sum = _sqlfunc.sum(StatTransaction.valeur)

    rows = (await db.execute(
        select(partner_col.label("pid"), StatTransaction.ressource.label("res"), _sum.label("v"))
        .where(W).group_by(partner_col, StatTransaction.ressource)
    )).all()

    data: dict = defaultdict(dict)   # pid -> {res: valeur}
    part_tot: dict = defaultdict(float)
    for r in rows:
        v = float(r.v or 0)
        if v <= 0:
            continue
        data[r.pid][r.res] = data[r.pid].get(r.res, 0.0) + v
        part_tot[r.pid] += v

    pids = sorted(part_tot, key=lambda p: part_tot[p], reverse=True)
    noms = {}
    if pids:
        rn = (await db.execute(select(RefPays.id, RefPays.nom_fr).where(RefPays.id.in_(pids)))).all()
        noms = {rid: nom for rid, nom in rn}
    codes = {c for d in data.values() for c in d}
    libs = {}
    if codes:
        rl = (await db.execute(select(StatRessource.nom_en, StatRessource.libelle).where(StatRessource.nom_en.in_(codes)))).all()
        libs = {code: lib for code, lib in rl}

    partenaires = []
    for pid in pids:
        lignes = sorted(
            ({"ressource": libs.get(c) or c, "valeur": v} for c, v in data[pid].items()),
            key=lambda x: x["valeur"], reverse=True,
        )
        partenaires.append({"nom": noms.get(pid) or "—", "total": part_tot[pid], "lignes": lignes})
    return {"partenaires": partenaires}


@router.get("/commerce/bilateral")
async def commerce_bilateral(
    db: AsyncSession = Depends(get_db),
    pays_a: int = Query(...),
    pays_b: int = Query(...),
):
    """Échanges bilatéraux cumulés entre deux pays (toutes ressources, toutes années),
    avec la ventilation par ressource dans chaque sens."""
    from sqlalchemy import and_ as _and, or_ as _or
    _sum = _sqlfunc.coalesce(_sqlfunc.sum(StatTransaction.valeur), 0)
    _s = _sqlfunc.sum(StatTransaction.valeur)

    ab = (await db.execute(select(_sum).where(
        StatTransaction.exportateur_id == pays_a, StatTransaction.importateur_id == pays_b))).scalar_one()
    ba = (await db.execute(select(_sum).where(
        StatTransaction.exportateur_id == pays_b, StatTransaction.importateur_id == pays_a))).scalar_one()
    bornes = (await db.execute(
        select(_sqlfunc.min(StatTransaction.annee), _sqlfunc.max(StatTransaction.annee)).where(
            _or(_and(StatTransaction.exportateur_id == pays_a, StatTransaction.importateur_id == pays_b),
                _and(StatTransaction.exportateur_id == pays_b, StatTransaction.importateur_id == pays_a)))
    )).first()

    async def totaux_import(imp_id):
        """Total des importations de imp_id (tous fournisseurs), global et par ressource."""
        rows = (await db.execute(
            select(StatTransaction.ressource, _s.label("v"))
            .where(StatTransaction.importateur_id == imp_id)
            .group_by(StatTransaction.ressource)
        )).all()
        par_res = {r.ressource: float(r.v or 0) for r in rows if (r.v or 0) > 0}
        return par_res, sum(par_res.values())

    b_imp, b_imp_tot = await totaux_import(pays_b)   # sens A→B : B importe
    a_imp, a_imp_tot = await totaux_import(pays_a)   # sens B→A : A importe

    async def ventil(exp_id, imp_id, imp_res):
        rows = (await db.execute(
            select(StatTransaction.ressource, _s.label("v"))
            .where(StatTransaction.exportateur_id == exp_id, StatTransaction.importateur_id == imp_id)
            .group_by(StatTransaction.ressource).order_by(_s.desc().nullslast())
        )).all()
        codes = [r.ressource for r in rows]
        libs = {}
        if codes:
            rl = (await db.execute(select(StatRessource.nom_en, StatRessource.libelle).where(StatRessource.nom_en.in_(codes)))).all()
            libs = {c: l for c, l in rl}
        out = []
        for r in rows:
            v = float(r.v or 0)
            if v <= 0:
                continue
            tot_res = imp_res.get(r.ressource, 0)
            out.append({
                "ressource": libs.get(r.ressource) or r.ressource,
                "valeur": v,
                "part_dependance": (v / tot_res) if tot_res > 0 else None,
            })
        return out

    # Groupements communs aux deux pays (ZLECAf, UA, OPEP…)
    from app.models.shared import RefGroupement, RefPaysGroupement
    communs_ids = (await db.execute(
        select(RefPaysGroupement.groupement_id)
        .where(RefPaysGroupement.pays_id.in_([pays_a, pays_b]))
        .group_by(RefPaysGroupement.groupement_id)
        .having(_sqlfunc.count(_sqlfunc.distinct(RefPaysGroupement.pays_id)) == 2)
    )).scalars().all()
    groupements = []
    if communs_ids:
        grows = (await db.execute(select(RefGroupement).where(RefGroupement.id.in_(communs_ids)).order_by(RefGroupement.nom_fr))).scalars().all()
        groupements = [{"code": g.code, "nom": g.nom_fr} for g in grows]

    # Accords / traités signés entre les deux pays
    from app.models.accord import Accord
    arows = (await db.execute(
        select(Accord).where(
            Accord.parties_pays_ids.contains([pays_a]),
            Accord.parties_pays_ids.contains([pays_b]),
            Accord.est_publie.is_(True),
            Accord.is_deleted.is_(False),
        ).order_by(Accord.date_signature.desc().nullslast())
    )).scalars().all()
    # Objet complet : la Fiche Pays ouvre le même modal que la page Accords
    accords = [{
        "id": a.id,
        "titre": a.titre,
        "reference": a.reference,
        "commentaires": a.commentaires,
        "date_signature": a.date_signature.isoformat() if a.date_signature else None,
        "date_entree_vigueur": a.date_entree_vigueur.isoformat() if a.date_entree_vigueur else None,
        "date_expiration": a.date_expiration.isoformat() if a.date_expiration else None,
        "statut": a.statut,
        "parties_pays_ids": a.parties_pays_ids or [],
        "parties_signataires": a.parties_signataires,
        "secteur_ids": a.secteur_ids or [],
        "branche_ids": a.branche_ids or [],
        "activite_ids": a.activite_ids or [],
    } for a in arows]

    return {
        "a_vers_b": float(ab or 0),
        "b_vers_a": float(ba or 0),
        "annee_min": bornes[0] if bornes else None,
        "annee_max": bornes[1] if bornes else None,
        # Dépendance globale : part du fournisseur dans le total des imports de l'importateur
        "a_vers_b_dependance": (float(ab or 0) / b_imp_tot) if b_imp_tot > 0 else None,
        "b_vers_a_dependance": (float(ba or 0) / a_imp_tot) if a_imp_tot > 0 else None,
        "a_vers_b_ressources": await ventil(pays_a, pays_b, b_imp),
        "b_vers_a_ressources": await ventil(pays_b, pays_a, a_imp),
        "groupements_communs": groupements,
        "accords": accords,
    }


@router.get("/commerce/transactions")
async def commerce_transactions(
    db: AsyncSession = Depends(get_db),
    annee: Optional[int] = Query(default=None),
    annee_min: Optional[int] = Query(default=None),
    annee_max: Optional[int] = Query(default=None),
    annees: Optional[str] = Query(default=None),
    exportateur_id: Optional[int] = Query(default=None),
    importateur_id: Optional[int] = Query(default=None),
    ressource: Optional[str] = Query(default=None),
    ressources: Optional[str] = Query(default=None),
    recherche: Optional[str] = Query(default=None),
    page: int = Query(default=1, ge=1),
    taille: int = Query(default=50, ge=1, le=200),
):
    """Tableau paginé public des flux commerciaux bilatéraux."""
    from sqlalchemy import or_ as _or
    from sqlalchemy.orm import aliased as _aliased

    Exp = _aliased(RefPays)
    Imp = _aliased(RefPays)

    base = (
        select(
            StatTransaction.id, StatTransaction.annee, StatTransaction.valeur,
            StatTransaction.ressource,
            Exp.nom_fr.label("exp_nom"), Imp.nom_fr.label("imp_nom"),
            StatTransaction.exportateur_id, StatTransaction.importateur_id,
            StatRessource.libelle.label("ressource_lib"),
        )
        .join(Exp, Exp.id == StatTransaction.exportateur_id)
        .join(Imp, Imp.id == StatTransaction.importateur_id)
        .outerjoin(StatRessource, StatRessource.nom_en == StatTransaction.ressource)
    )
    if annee is not None:
        base = base.where(StatTransaction.annee == annee)
    if annee_min is not None:
        base = base.where(StatTransaction.annee >= annee_min)
    if annee_max is not None:
        base = base.where(StatTransaction.annee <= annee_max)
    if annees:
        liste_a = [int(x) for x in annees.split(",") if x.strip().isdigit()]
        if liste_a:
            base = base.where(StatTransaction.annee.in_(liste_a))
    if exportateur_id is not None:
        base = base.where(StatTransaction.exportateur_id == exportateur_id)
    if importateur_id is not None:
        base = base.where(StatTransaction.importateur_id == importateur_id)
    if ressource:
        base = base.where(StatTransaction.ressource == ressource)
    if ressources:
        liste_r = [x for x in ressources.split(",") if x.strip()]
        if liste_r:
            base = base.where(StatTransaction.ressource.in_(liste_r))
    if recherche:
        motif = f"%{recherche.strip()}%"
        base = base.where(_or(Exp.nom_fr.ilike(motif), Imp.nom_fr.ilike(motif),
                              StatTransaction.ressource.ilike(motif),
                              StatRessource.libelle.ilike(motif)))

    total = (await db.execute(select(_sqlfunc.count()).select_from(base.subquery()))).scalar_one()
    rows = (await db.execute(
        base.order_by(StatTransaction.valeur.desc().nullslast(), StatTransaction.annee.desc())
            .offset((page - 1) * taille).limit(taille)
    )).all()
    return {
        "total": total, "page": page, "taille": taille,
        "lignes": [
            {
                "id": r.id, "annee": r.annee,
                "exportateur": r.exp_nom, "importateur": r.imp_nom,
                "exportateur_id": r.exportateur_id, "importateur_id": r.importateur_id,
                "ressource": r.ressource_lib or r.ressource,
                "valeur": float(r.valeur) if r.valeur is not None else None,
            }
            for r in rows
        ],
    }


@router.get("/commerce/kpis")
async def commerce_kpis(
    db: AsyncSession = Depends(get_db),
    pays_id: int = Query(...),
    direction: str = Query(default="exportateur"),   # exportateur | importateur
    annee_min: Optional[int] = Query(default=None),
    annee_max: Optional[int] = Query(default=None),
    annees: Optional[str] = Query(default=None),
    ressources: Optional[str] = Query(default=None),
):
    """Indicateurs agrégés des flux d'un pays (vue exportateur ou importateur)."""
    from sqlalchemy import and_ as _and

    self_col = StatTransaction.exportateur_id if direction == "exportateur" else StatTransaction.importateur_id
    partner_col = StatTransaction.importateur_id if direction == "exportateur" else StatTransaction.exportateur_id

    conds = [self_col == pays_id]
    if annee_min is not None:
        conds.append(StatTransaction.annee >= annee_min)
    if annee_max is not None:
        conds.append(StatTransaction.annee <= annee_max)
    if annees:
        la = [int(x) for x in annees.split(",") if x.strip().isdigit()]
        if la:
            conds.append(StatTransaction.annee.in_(la))
    if ressources:
        lr = [x for x in ressources.split(",") if x.strip()]
        if lr:
            conds.append(StatTransaction.ressource.in_(lr))
    W = _and(*conds)
    # Conditions hors période : pour comparer une année précise (n) à n-1 sans
    # être bridé par annee_min/annee_max.
    base_conds = [self_col == pays_id]
    if ressources:
        lr0 = [x for x in ressources.split(",") if x.strip()]
        if lr0:
            base_conds.append(StatTransaction.ressource.in_(lr0))
    _sum = _sqlfunc.sum(StatTransaction.valeur)

    # « Année record » = année au plus fort volume total, sur toute la période filtrée.
    rec = (await db.execute(
        select(StatTransaction.annee, _sum.label("v")).where(W)
        .group_by(StatTransaction.annee).order_by(_sum.desc().nullslast()).limit(1)
    )).first()
    annee_record = {"annee": rec.annee, "valeur": float(rec.v or 0)} if rec else None

    # Tous les autres indicateurs portent sur la DERNIÈRE année sélectionnée
    # (ex. filtre 2021–2024 → valeurs de 2024), calculés année par année.
    annee_ref = (await db.execute(select(_sqlfunc.max(StatTransaction.annee)).where(W))).scalar_one()

    total = 0.0
    top_partenaire = None
    top_ressource = None
    nb_partenaires = 0
    part_top = None

    if annee_ref is not None:
        Wref = _and(W, StatTransaction.annee == annee_ref)

        total = float((await db.execute(select(_sqlfunc.coalesce(_sum, 0)).where(Wref))).scalar_one() or 0)

        tp = (await db.execute(
            select(partner_col.label("pid"), _sum.label("v")).where(Wref)
            .group_by(partner_col).order_by(_sum.desc().nullslast()).limit(1)
        )).first()
        if tp:
            nom = (await db.execute(select(RefPays.nom_fr).where(RefPays.id == tp.pid))).scalar_one_or_none()
            top_partenaire = {"id": tp.pid, "nom": nom, "valeur": float(tp.v or 0)}

        tr = (await db.execute(
            select(StatTransaction.ressource, _sum.label("v")).where(Wref)
            .group_by(StatTransaction.ressource).order_by(_sum.desc().nullslast()).limit(1)
        )).first()
        if tr:
            lib = (await db.execute(select(StatRessource.libelle).where(StatRessource.nom_en == tr.ressource))).scalar_one_or_none()
            top_ressource = {"ressource": lib or tr.ressource, "valeur": float(tr.v or 0)}

        nb_partenaires = int((await db.execute(select(_sqlfunc.count(_sqlfunc.distinct(partner_col))).where(Wref))).scalar_one() or 0)
        part_top = (top_partenaire["valeur"] / total * 100) if (top_partenaire and total > 0) else None

    # Variation n vs n-1 en VERROUILLANT le 1er partenaire de l'année n : on
    # prend la valeur (et la part) de CE partenaire à n-1, peu importe qui était
    # 1er cette année-là.
    part_top_variation = None
    annee_prec_ref = None
    if annee_ref is not None and top_partenaire:
        annee_prec = annee_ref - 1
        Wprec = _and(*base_conds, StatTransaction.annee == annee_prec)
        vp = float((await db.execute(
            select(_sqlfunc.coalesce(_sum, 0)).where(_and(Wprec, partner_col == top_partenaire["id"]))
        )).scalar_one() or 0)
        total_prec = float((await db.execute(select(_sqlfunc.coalesce(_sum, 0)).where(Wprec))).scalar_one() or 0)
        part_prec = (vp / total_prec * 100) if total_prec > 0 else None
        top_partenaire["valeur_prec"] = vp
        top_partenaire["variation"] = ((top_partenaire["valeur"] - vp) / abs(vp) * 100) if vp > 0 else None
        if vp > 0 or (part_prec is not None):
            annee_prec_ref = annee_prec
        top_partenaire["annee_prec"] = annee_prec_ref
        if part_top is not None and part_prec:
            part_top_variation = (part_top - part_prec) / abs(part_prec) * 100

    return {
        "annee_ref": annee_ref,
        "total": total,
        "annee_record": annee_record,
        "top_partenaire": top_partenaire,
        "top_ressource": top_ressource,
        "nb_partenaires": nb_partenaires,
        "part_top_partenaire": part_top,
        "part_top_partenaire_variation": part_top_variation,
        "annee_prec": annee_prec_ref,
    }


@router.get("/commerce/tops")
async def commerce_tops(
    db: AsyncSession = Depends(get_db),
    pays_id: int = Query(...),
    direction: str = Query(default="exportateur"),
    annee_min: Optional[int] = Query(default=None),
    annee_max: Optional[int] = Query(default=None),
    annees: Optional[str] = Query(default=None),
    ressources: Optional[str] = Query(default=None),
    limite: int = Query(default=10, ge=1, le=50),
):
    """Top partenaires et top ressources d'un pays, en cumul sur la période filtrée."""
    from sqlalchemy import and_ as _and

    self_col = StatTransaction.exportateur_id if direction == "exportateur" else StatTransaction.importateur_id
    partner_col = StatTransaction.importateur_id if direction == "exportateur" else StatTransaction.exportateur_id

    conds = [self_col == pays_id]
    if annee_min is not None:
        conds.append(StatTransaction.annee >= annee_min)
    if annee_max is not None:
        conds.append(StatTransaction.annee <= annee_max)
    if annees:
        la = [int(x) for x in annees.split(",") if x.strip().isdigit()]
        if la:
            conds.append(StatTransaction.annee.in_(la))
    if ressources:
        lr = [x for x in ressources.split(",") if x.strip()]
        if lr:
            conds.append(StatTransaction.ressource.in_(lr))
    W = _and(*conds)
    _sum = _sqlfunc.sum(StatTransaction.valeur)

    prows = (await db.execute(
        select(partner_col.label("pid"), _sum.label("v")).where(W)
        .group_by(partner_col).order_by(_sum.desc().nullslast()).limit(limite)
    )).all()
    noms = {}
    iso2 = {}
    pids = [r.pid for r in prows]
    if pids:
        rn = (await db.execute(select(RefPays.id, RefPays.nom_fr, RefPays.code_iso2).where(RefPays.id.in_(pids)))).all()
        noms = {rid: nom for rid, nom, _ in rn}
        iso2 = {rid: c for rid, _, c in rn}
    partenaires = [{"id": r.pid, "nom": noms.get(r.pid) or "—", "valeur": float(r.v or 0), "code_iso2": iso2.get(r.pid)} for r in prows]

    rrows = (await db.execute(
        select(StatTransaction.ressource.label("res"), _sum.label("v")).where(W)
        .group_by(StatTransaction.ressource).order_by(_sum.desc().nullslast()).limit(limite)
    )).all()
    libs = {}
    codes = [r.res for r in rrows]
    if codes:
        rl = (await db.execute(select(StatRessource.nom_en, StatRessource.libelle).where(StatRessource.nom_en.in_(codes)))).all()
        libs = {code: lib for code, lib in rl}
    ressources_top = [{"ressource": libs.get(r.res) or r.res, "valeur": float(r.v or 0)} for r in rrows]

    total = float((await db.execute(select(_sqlfunc.coalesce(_sum, 0)).where(W))).scalar_one() or 0)

    return {"partenaires": partenaires, "ressources": ressources_top, "total": total}


@router.get("/commerce/concentration")
async def commerce_concentration(
    db: AsyncSession = Depends(get_db),
    pays_id: int = Query(...),
    direction: str = Query(default="exportateur"),
    annee_min: Optional[int] = Query(default=None),
    annee_max: Optional[int] = Query(default=None),
    annees: Optional[str] = Query(default=None),
    ressources: Optional[str] = Query(default=None),
    limite: int = Query(default=40, ge=1, le=200),
):
    """Courbe de concentration (Pareto) : part cumulée des débouchés classés."""
    from sqlalchemy import and_ as _and

    self_col = StatTransaction.exportateur_id if direction == "exportateur" else StatTransaction.importateur_id
    partner_col = StatTransaction.importateur_id if direction == "exportateur" else StatTransaction.exportateur_id

    conds = [self_col == pays_id]
    if annee_min is not None:
        conds.append(StatTransaction.annee >= annee_min)
    if annee_max is not None:
        conds.append(StatTransaction.annee <= annee_max)
    if annees:
        la = [int(x) for x in annees.split(",") if x.strip().isdigit()]
        if la:
            conds.append(StatTransaction.annee.in_(la))
    if ressources:
        lr = [x for x in ressources.split(",") if x.strip()]
        if lr:
            conds.append(StatTransaction.ressource.in_(lr))
    W = _and(*conds)
    _sum = _sqlfunc.sum(StatTransaction.valeur)

    rows = (await db.execute(
        select(partner_col.label("pid"), _sum.label("v")).where(W)
        .group_by(partner_col).order_by(_sum.desc().nullslast())
    )).all()
    total = sum(float(r.v or 0) for r in rows)
    top = rows[:limite]
    noms = {}
    pids = [r.pid for r in top]
    if pids:
        rn = (await db.execute(select(RefPays.id, RefPays.nom_fr).where(RefPays.id.in_(pids)))).all()
        noms = {rid: nom for rid, nom in rn}

    points = []
    cumul = 0.0
    for i, r in enumerate(top, start=1):
        v = float(r.v or 0)
        cumul += v
        points.append({
            "rang": i,
            "nom": noms.get(r.pid) or "—",
            "part": (v / total * 100) if total > 0 else 0,
            "part_cumulee": (cumul / total * 100) if total > 0 else 0,
        })
    return {"points": points, "total_partenaires": len(rows)}


@router.get("/commerce/repartition")
async def commerce_repartition(
    db: AsyncSession = Depends(get_db),
    pays_id: int = Query(...),
    direction: str = Query(default="exportateur"),
    annee_min: Optional[int] = Query(default=None),
    annee_max: Optional[int] = Query(default=None),
    annees: Optional[str] = Query(default=None),
    ressources: Optional[str] = Query(default=None),
    limite: int = Query(default=10, ge=1, le=50),
):
    """Ventilation par partenaire ET par ressource (barres empilées)."""
    from sqlalchemy import and_ as _and
    from collections import defaultdict

    self_col = StatTransaction.exportateur_id if direction == "exportateur" else StatTransaction.importateur_id
    partner_col = StatTransaction.importateur_id if direction == "exportateur" else StatTransaction.exportateur_id

    conds = [self_col == pays_id]
    if annee_min is not None:
        conds.append(StatTransaction.annee >= annee_min)
    if annee_max is not None:
        conds.append(StatTransaction.annee <= annee_max)
    if annees:
        la = [int(x) for x in annees.split(",") if x.strip().isdigit()]
        if la:
            conds.append(StatTransaction.annee.in_(la))
    if ressources:
        lr = [x for x in ressources.split(",") if x.strip()]
        if lr:
            conds.append(StatTransaction.ressource.in_(lr))
    W = _and(*conds)
    _sum = _sqlfunc.sum(StatTransaction.valeur)

    rows = (await db.execute(
        select(partner_col.label("pid"), StatTransaction.ressource.label("res"), _sum.label("v"))
        .where(W).group_by(partner_col, StatTransaction.ressource)
    )).all()

    part_tot: dict = defaultdict(float)
    res_tot: dict = defaultdict(float)
    cell: dict = defaultdict(float)
    for r in rows:
        v = float(r.v or 0)
        if v <= 0:
            continue
        part_tot[r.pid] += v
        res_tot[r.res] += v
        cell[(r.pid, r.res)] += v

    top_pids = sorted(part_tot, key=lambda p: part_tot[p], reverse=True)[:limite]
    res_order = sorted(res_tot, key=lambda r: res_tot[r], reverse=True)

    libs = {}
    if res_order:
        rl = (await db.execute(select(StatRessource.nom_en, StatRessource.libelle).where(StatRessource.nom_en.in_(res_order)))).all()
        libs = {code: lib for code, lib in rl}
    noms = {}
    if top_pids:
        rn = (await db.execute(select(RefPays.id, RefPays.nom_fr).where(RefPays.id.in_(top_pids)))).all()
        noms = {rid: nom for rid, nom in rn}

    ressources_labels = [libs.get(r) or r for r in res_order]
    partenaires = [
        {
            "nom": noms.get(pid) or "—",
            "total": part_tot[pid],
            "valeurs": [cell.get((pid, res), 0.0) for res in res_order],
        }
        for pid in top_pids
    ]
    return {"ressources": ressources_labels, "partenaires": partenaires}


@router.get("/commerce/balance")
async def commerce_balance(
    db: AsyncSession = Depends(get_db),
    pays_id: int = Query(...),
    annee_min: Optional[int] = Query(default=None),
    annee_max: Optional[int] = Query(default=None),
    annees: Optional[str] = Query(default=None),
    ressources: Optional[str] = Query(default=None),
):
    """Balance commerciale annuelle d'un pays : exportations − importations."""
    from sqlalchemy import and_ as _and

    def periode_conds():
        c = []
        if annee_min is not None:
            c.append(StatTransaction.annee >= annee_min)
        if annee_max is not None:
            c.append(StatTransaction.annee <= annee_max)
        if annees:
            la = [int(x) for x in annees.split(",") if x.strip().isdigit()]
            if la:
                c.append(StatTransaction.annee.in_(la))
        if ressources:
            lr = [x for x in ressources.split(",") if x.strip()]
            if lr:
                c.append(StatTransaction.ressource.in_(lr))
        return c

    _sum = _sqlfunc.sum(StatTransaction.valeur)

    async def par_annee(col):
        rows = (await db.execute(
            select(StatTransaction.annee, _sum.label("v"))
            .where(_and(col == pays_id, *periode_conds()))
            .group_by(StatTransaction.annee)
        )).all()
        return {r.annee: float(r.v or 0) for r in rows}

    exp = await par_annee(StatTransaction.exportateur_id)
    imp = await par_annee(StatTransaction.importateur_id)

    annees_all = sorted(set(exp) | set(imp))
    return [
        {
            "annee": a,
            "exportations": exp.get(a, 0.0),
            "importations": imp.get(a, 0.0),
            "balance": exp.get(a, 0.0) - imp.get(a, 0.0),
        }
        for a in annees_all
    ]


# ── GET /statistiques/entreprises-siege ──────────────────────────────────────
# Fiche Pays (Sénégal × pays X) : entreprises installées au Sénégal dont le
# siège est dans le pays X (table entreprises_installees, publiées uniquement).
@router.get("/entreprises-siege")
async def entreprises_par_siege(
    pays_id: int = Query(...),
    db: AsyncSession = Depends(get_db),
):
    from app.models.entreprise import EntrepriseIntallee, RefSecteur

    rows = (await db.execute(
        select(EntrepriseIntallee)
        .where(
            EntrepriseIntallee.siege_pays_id == pays_id,
            EntrepriseIntallee.is_deleted == False,  # noqa: E712
            EntrepriseIntallee.est_publie == True,   # noqa: E712
        )
        .order_by(EntrepriseIntallee.nom)
    )).scalars().all()

    # Noms de secteurs pour l'affichage compact (les ids sont stockés en ARRAY)
    sect_ids = {i for e in rows for i in (e.secteur_ids or [])}
    secteurs: dict[int, str] = {}
    if sect_ids:
        secteurs = {s.id: s.nom for s in (await db.execute(
            select(RefSecteur).where(RefSecteur.id.in_(sect_ids)))).scalars().all()}

    return {
        "total": len(rows),
        "entreprises": [{
            "id": e.id,
            "nom": e.nom,
            "forme_juridique": e.forme_juridique,
            "region": e.region.nom if e.region else None,
            "secteurs": [secteurs[i] for i in (e.secteur_ids or []) if i in secteurs],
        } for e in rows],
    }
