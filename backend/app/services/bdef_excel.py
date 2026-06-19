"""
Parseur des fichiers Excel BDEF (source ANSD).

Voir docs/bdef_structure_import.md pour la structure complète du fichier.

Deux niveaux :
  - `extraire_blocs(rows, feuille)` : fonction PURE qui découpe une liste de
    lignes brutes (list[tuple]) en blocs TABLEAU. Testable sans fichier.
  - `parser_fichier(path)` : charge le classeur openpyxl et applique
    `extraire_blocs` aux feuilles COMPTES et RATIOS.

Le parseur extrait TOUTES les valeurs disponibles (tous les codes Réf., toutes
les paires de ratios, tous les libellés D/E). La sélection des valeurs utiles et
le calcul des indicateurs sont la responsabilité de la couche d'import, pas du
parseur — ce qui le rend robuste aux évolutions de la nomenclature.
"""
import re
from dataclasses import dataclass, field

from app.utils.bdef_matching import normaliser, detecter_niveau

FEUILLE_COMPTES = "COMPTES"
FEUILLE_RATIOS  = "RATIOS"

# Types de blocs
#   COMPTES : A → ACTIF/PASSIF (bilan), B → CHARGES/PRODUITS (compte de résultat)
#   RATIOS  : C → ratios (modèle fraction), D → CAF/trésorerie, E → autres indicateurs
_RE_TABLEAU = re.compile(r"TABLEAU\s*:\s*(\d+)\s*-\s*([A-E])\b\s*([A-ZÀ-Ÿ]+)?", re.I)

_AN_MIN, _AN_MAX = 1990, 2100


def _txt(cell) -> str:
    return str(cell).strip() if cell is not None else ""


def _num(cell):
    """Retourne un float si la cellule est numérique, sinon None."""
    if isinstance(cell, bool):
        return None
    if isinstance(cell, (int, float)):
        return float(cell)
    return None


@dataclass
class LigneValeur:
    """Une ligne de données extraite, avec ses valeurs par année."""
    cle: str                       # 'ref:TI' | 'lib:production' | 'ratio:<num>||<den>'
    libelle_brut: str              # libellé original (traçabilité)
    valeurs: dict[int, float] = field(default_factory=dict)   # année → valeur


@dataclass
class BlocTableau:
    code: str                      # code BDEF brut ("0", "3", "101", "201")
    niveau: str | None             # global | secteur | groupe | macro_secteur
    libelle_secteur: str           # nom brut du secteur (pour le matching)
    feuille: str                   # COMPTES | RATIOS
    type_bloc: str                 # A | B | C | D | E
    sous_type: str                 # ACTIF | PASSIF | CHARGES | PRODUITS | ''
    annees: list[int] = field(default_factory=list)
    lignes: list[LigneValeur] = field(default_factory=list)


def _colonnes_annees(rows: list, debut: int, fin: int) -> dict[int, int]:
    """
    Repère la ligne d'en-tête des années dans [debut, fin) et retourne
    {index_colonne: année}. Une ligne d'années a au moins 3 entiers à 4 chiffres
    plausibles. Ne présume aucune colonne fixe (cf. doc §4).
    """
    for r in range(debut, fin):
        trouve = {}
        for ci, cell in enumerate(rows[r]):
            v = _num(cell)
            if v is not None and v == int(v) and _AN_MIN <= int(v) <= _AN_MAX:
                trouve[ci] = int(v)
        if len(trouve) >= 3:
            return trouve
    return {}


def _carte_codes_noms(rows: list) -> dict[str, str]:
    """
    Première passe : { code → nom du secteur } depuis les lignes
    « <code> | - | <NOM> ». Le nom n'apparaît pas dans chaque bloc (notamment
    les tableaux D/E des RATIOS) ; on le récupère donc globalement.
    """
    carte = {}
    for row in rows:
        if len(row) < 3:
            continue
        code, tiret, nom = _txt(row[0]), _txt(row[1]), _txt(row[2])
        if tiret == "-" and code.isdigit() and nom and code not in carte:
            carte[code] = nom
    return carte


def _bornes_blocs(rows: list) -> list[tuple[int, re.Match]]:
    """Retourne [(index_ligne, match_entête)] pour chaque en-tête TABLEAU."""
    bornes = []
    for i, row in enumerate(rows):
        a = _txt(row[0]) if row else ""
        m = _RE_TABLEAU.search(a)
        if m:
            bornes.append((i, m))
    return bornes


def _extraire_comptes(rows, debut, fin, cols) -> list[LigneValeur]:
    """COMPTES : lecture par code Réf. (colonne B), valeurs sous les années."""
    lignes = []
    for r in range(debut, fin):
        row = rows[r]
        ref = _txt(row[1]) if len(row) > 1 else ""
        if not ref or ref.lower() == "réf." or not re.fullmatch(r"[A-Z]{2}", ref):
            continue
        valeurs = {}
        for ci, annee in cols.items():
            if ci < len(row):
                v = _num(row[ci])
                if v is not None:
                    valeurs[annee] = v
        if valeurs:
            lignes.append(LigneValeur(cle=f"ref:{ref}", libelle_brut=_txt(row[0]), valeurs=valeurs))
    return lignes


def _extraire_simple(rows, debut, fin, cols) -> list[LigneValeur]:
    """RATIOS D/E : libellé simple en colonne A, valeurs sous les années."""
    lignes = []
    for r in range(debut, fin):
        row = rows[r]
        lib = _txt(row[0]) if row else ""
        if not lib:
            continue
        valeurs = {}
        for ci, annee in cols.items():
            if ci < len(row):
                v = _num(row[ci])
                if v is not None:
                    valeurs[annee] = v
        if valeurs:
            lignes.append(LigneValeur(cle=f"lib:{normaliser(lib)}", libelle_brut=lib, valeurs=valeurs))
    return lignes


def _extraire_ratios(rows, debut, fin, cols) -> list[LigneValeur]:
    """
    RATIOS C : modèle fraction. Le numérateur porte le libellé + les valeurs ;
    le dénominateur est le libellé de la (des) ligne(s) suivante(s) sans valeur.
    La clé du ratio est 'ratio:<num_normalisé>||<den_normalisé>'.
    """
    lignes = []
    r = debut
    while r < fin:
        row = rows[r]
        lib = _txt(row[0]) if row else ""
        if not lib or normaliser(lib) == "ratios":
            r += 1
            continue
        valeurs = {}
        for ci, annee in cols.items():
            if ci < len(row):
                v = _num(row[ci])
                if v is not None:
                    valeurs[annee] = v
        if not valeurs:
            r += 1
            continue
        # numérateur trouvé → chercher le dénominateur (prochaine ligne avec
        # libellé mais sans aucune valeur sous les colonnes années)
        num_lib = lib
        den_lib = ""
        rr = r + 1
        while rr < fin:
            nrow = rows[rr]
            nlib = _txt(nrow[0]) if nrow else ""
            nval = any(
                _num(nrow[ci]) is not None for ci in cols if ci < len(nrow)
            )
            if nlib and not nval:
                den_lib = nlib
                break
            if nlib and nval:  # ligne = numérateur suivant, pas de dénominateur isolé
                break
            rr += 1
        cle = f"ratio:{normaliser(num_lib)}||{normaliser(den_lib)}"
        lignes.append(LigneValeur(cle=cle, libelle_brut=f"{num_lib} / {den_lib}".strip(" /"), valeurs=valeurs))
        r = rr if den_lib else r + 1
    return lignes


def extraire_blocs(rows: list, feuille: str) -> list[BlocTableau]:
    """
    Découpe une feuille (liste de lignes brutes) en blocs TABLEAU.

    `rows`    : list de tuples/listes (valeurs des cellules), 0-indexée.
    `feuille` : FEUILLE_COMPTES ou FEUILLE_RATIOS.
    """
    if not rows:
        return []
    carte_noms = _carte_codes_noms(rows)
    bornes = _bornes_blocs(rows)
    blocs = []
    for idx, (ligne_debut, m) in enumerate(bornes):
        code = m.group(1)
        type_bloc = m.group(2).upper()
        sous_type = (m.group(3) or "").upper()
        fin = bornes[idx + 1][0] if idx + 1 < len(bornes) else len(rows)
        cols = _colonnes_annees(rows, ligne_debut, fin)
        if not cols:
            continue
        if feuille == FEUILLE_COMPTES:
            lignes = _extraire_comptes(rows, ligne_debut, fin, cols)
        elif type_bloc == "C":
            lignes = _extraire_ratios(rows, ligne_debut, fin, cols)
        else:  # D, E
            lignes = _extraire_simple(rows, ligne_debut, fin, cols)
        if not lignes:
            continue
        blocs.append(BlocTableau(
            code=code,
            niveau=detecter_niveau(code),
            libelle_secteur=carte_noms.get(code, ""),
            feuille=feuille,
            type_bloc=type_bloc,
            sous_type=sous_type,
            annees=sorted(set(cols.values())),
            lignes=lignes,
        ))
    return blocs


def parser_fichier(path: str) -> dict[str, list[BlocTableau]]:
    """
    Charge un classeur BDEF et retourne {feuille: [BlocTableau]} pour les
    feuilles COMPTES et RATIOS. La feuille Analyse (dérivée) est ignorée.
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    resultat = {}
    correspondances = {
        FEUILLE_COMPTES: "EDITIONS COMPTES",
        FEUILLE_RATIOS:  "EDITIONS RATIOS",
    }
    for cle, titre in correspondances.items():
        if titre not in wb.sheetnames:
            continue
        ws = wb[titre]
        rows = list(ws.iter_rows(values_only=True))
        resultat[cle] = extraire_blocs(rows, cle)
    wb.close()
    return resultat
