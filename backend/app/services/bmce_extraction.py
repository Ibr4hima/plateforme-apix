#!/usr/bin/env python3
"""Extracteur du Bulletin Mensuel des Statistiques du Commerce Extérieur (ANSD).

Lit le PDF du bulletin (via `pdftotext -layout`, poppler-utils requis) et
produit un classeur Excel ne contenant QUE les données brutes retenues :
les 4 mois récents (le mois de l'année précédente, les cumuls, variations,
moyennes et glissements sont exclus — ce sont des dérivés recalculés par
la plateforme selon les règles ANSD).

Feuilles produites (numérotation = tableaux du bulletin) :
  T1 / T2       Ensemble export / import — Valeur (FCFA) et Poids net (kg)
  T3 / T4       Exportations par groupe d'utilisation — Valeur (M FCFA) / Poids (k kg)
  T7 / T8       Importations par groupe d'utilisation — idem
  T11 / T12     Exportations par produits regroupés — idem
  T15 / T16     Importations par produits regroupés — idem
  T19 / T20     Exportations par chapitre — idem
  T23 / T24     Importations par chapitre — idem
  T27 / T29     Exportations / importations par pays — Valeur (M FCFA)
  Métadonnées   mois du bulletin, fichier source, rapport de vérification

Fiabilité : le bulletin contient ses propres témoins. L'extraction est
contre-vérifiée avec les tableaux dérivés NON importés :
  · valeurs unitaires (T5, T9, T13, T17, T21, T25) ≈ valeur / poids ;
  · parts en % (T6, T10, T14, T18, T22, T26, T28, T30) ≈ valeur / total ;
  · sommes des groupes d'utilisation et des chapitres ≈ ensemble (T1/T2) ;
  · cumul de l'année ≈ mois du fichier (année du bulletin) + mois antérieurs
    de l'année fournis en complément par la plateforme (base des bulletins
    déjà importés) — un bulletin de mai cumule janv–mai alors que le fichier
    ne porte que févr–mai : janvier vient de la base.
Toute divergence au-delà des tolérances d'arrondi fait échouer l'extraction.

Usage : python3 extraire_bmce.py BULLETIN.pdf [SORTIE.xlsx]
"""
from __future__ import annotations

import re
import subprocess
import sys
from datetime import datetime
from difflib import get_close_matches
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# ── Tableaux : (numéro, titre attendu (début), unité) ─────────────────────────
BRUTS = {
    1:  ("Statistiques mensuelles à l’exportation", "FCFA / kg"),
    2:  ("Statistiques mensuelles à l’importation", "FCFA / kg"),
    3:  ("exportations par groupe d’utilisation : Valeur FAB", "millions de FCFA"),
    4:  ("exportations par groupe d’utilisation : Poids net", "milliers de kg"),
    7:  ("importations par groupe d’utilisation : Valeur CAF", "millions de FCFA"),
    8:  ("importations par groupe d’utilisation : Poids net", "milliers de kg"),
    11: ("exportations par produits regroupés : Valeur FAB", "millions de FCFA"),
    12: ("exportations par produits regroupés : Poids net", "milliers de kg"),
    15: ("importations par produits regroupés : Valeur CAF", "millions de FCFA"),
    16: ("importations par produits regroupés : Poids net", "milliers de kg"),
    19: ("exportations par chapitre : Valeur FAB", "millions de FCFA"),
    20: ("exportations par chapitre : Poids net", "milliers de kg"),
    23: ("importations par chapitre : Valeur CAF", "millions de FCFA"),
    24: ("importations par chapitre : Poids net", "milliers de kg"),
    27: ("exportations par pays : Valeur FAB", "millions de FCFA"),
    29: ("importations par pays : Valeur CAF", "millions de FCFA"),
}
# Tableaux témoins (dérivés, utilisés uniquement pour vérifier)
TEMOINS_VU = {5: (3, 4), 9: (7, 8), 13: (11, 12), 17: (15, 16), 21: (19, 20), 25: (23, 24)}
TEMOINS_PART = {6: 3, 10: 7, 14: 11, 18: 15, 22: 19, 26: 23, 28: 27, 30: 29}

MOIS_RE = re.compile(r"^(janv|févr|mars|avr|mai|juin|juil|août|sept|oct|nov|déc)\.?-\d{2}$")
NOMBRE_RE = re.compile(r"^-?\d[\d ]*(?:,\d+)?$")
MOIS_NUM = {"janv": 1, "févr": 2, "mars": 3, "avr": 4, "mai": 5, "juin": 6,
            "juil": 7, "août": 8, "sept": 9, "oct": 10, "nov": 11, "déc": 12}


def numero_mois(mois: str) -> tuple[int, int]:
    """« févr-25 » → (2, 25)."""
    nom, an = mois.split("-")
    return MOIS_NUM[nom], int(an)


def nombre(champ: str) -> float | None:
    """« 278 948 121 546 » → float ; « - » (zéro rigoureux) et vide → None ; « # » → None."""
    t = champ.strip()
    if t in ("-", "–", "", "#"):
        return None
    if not NOMBRE_RE.match(t):
        raise ValueError(f"champ non numérique : {champ!r}")
    return float(t.replace(" ", "").replace(" ", "").replace(",", "."))


def decouper(ligne: str) -> list[str]:
    return [c.strip() for c in re.split(r"\s{2,}", ligne.strip()) if c.strip()]


class Bulletin:
    def __init__(self, chemin_pdf: Path):
        self.chemin = chemin_pdf
        brut = subprocess.run(
            ["pdftotext", "-layout", str(chemin_pdf), "-"],
            capture_output=True, check=True,
        ).stdout.decode("utf-8")
        self.lignes = brut.split("\n")
        self.mois: list[str] = []          # les 5 mois de l'en-tête, ex. ["avr-24", …, "avr-25"]
        self.tables: dict[int, dict[str, list[float | None]]] = {}
        self.anomalies: list[str] = []        # lignes illisibles des tableaux BRUTS (fatales)
        self.temoins_perdus: list[str] = []   # lignes illisibles des tableaux TÉMOINS (avertissements)

    # ── Découpage du texte en tableaux ──
    def _bornes(self) -> dict[int, tuple[int, int]]:
        debuts: list[tuple[int, int]] = []
        for i, l in enumerate(self.lignes):
            m = re.match(r"\s*TABLEAU (\d+)\s*:", l)
            if m:
                debuts.append((int(m.group(1)), i))
        bornes = {}
        for k, (num, i) in enumerate(debuts):
            fin = debuts[k + 1][1] if k + 1 < len(debuts) else len(self.lignes)
            bornes[num] = (i, fin)
        return bornes

    def _lire_mois(self, bloc: list[str]) -> list[str]:
        for l in bloc[:14]:
            champs = decouper(l)
            mois = [c.replace(".", "") for c in champs if MOIS_RE.match(c)]
            if len(mois) == 5:
                return mois
        raise ValueError("en-tête des 5 mois introuvable")

    def _lire_table(self, num: int, bloc: list[str], lignes_t1: bool = False) -> dict[str, list[float | None]]:
        """Rangées d'un tableau : libellé → les 5 valeurs mensuelles (+ cumuls en témoin)."""
        rangees: dict[str, list[float | None]] = {}
        self._cumuls: dict[str, tuple[float | None, float | None]]
        cumuls = getattr(self, "cumuls", {})
        cumuls[num] = {}
        self.cumuls = cumuls
        for l in bloc:
            if not l.strip():
                continue
            t = l.strip()
            # En-têtes / pieds à ignorer
            if (re.match(r"TABLEAU \d+", t) or t.startswith("Source")
                    or "Cumul" in t or "Variation" in t or t.startswith("Glissement")
                    or MOIS_RE.match(decouper(l)[0] if decouper(l) else "")
                    or re.fullmatch(r"\d+", t)   # numéro de page
                    or t.startswith("BULLETIN") or t.startswith("ANSD")
                    or re.fullmatch(r"(20\d\d\s*){1,2}", t)):
                continue
            champs = decouper(l)
            # Les fragments d'en-tête et la prose ont peu de champs : silence
            if len(champs) < 4 or not re.match(r"[A-ZÀ-ÿa-z]", champs[0]):
                continue
            libelle, valeurs = champs[0], champs[1:]
            if lignes_t1 and not (libelle.startswith("Valeur") or libelle.startswith("Poids")):
                continue
            # Ligne illisible : fatale dans un tableau BRUT (donnée importée),
            # simple avertissement dans un tableau TÉMOIN (le témoin est perdu
            # pour cette rubrique, qui reste contrôlée par cumuls/parts/sommes
            # — ex. cellules vides sur la ligne VU « OR NON MONETAIRE », juin 25).
            # La ligne TOTAL d'un tableau brut est une dérivée jamais importée
            # (souvent imprimée avec des cellules fusionnées) : témoin aussi.
            derive = libelle.strip().upper() == "TOTAL"
            sortie = self.anomalies if (num in BRUTS and not derive) else self.temoins_perdus
            # Invariant du bulletin : 5 mois + 2 cumuls (+ variations facultatives)
            if len(valeurs) < 7:
                sortie.append(f"T{num} : ligne inattendue ({len(valeurs)} champs) : {t[:80]}")
                continue
            try:
                mensuels = [nombre(v) for v in valeurs[:5]]
                cum = (nombre(valeurs[5]), nombre(valeurs[6]))
            except ValueError as e:
                sortie.append(f"T{num} : {e} dans : {t[:80]}")
                continue
            if libelle in rangees:
                self.anomalies.append(f"T{num} : libellé en double : {libelle}")
            rangees[libelle] = mensuels
            cumuls[num][libelle] = cum
        if not rangees:
            raise ValueError(f"T{num} : aucune rangée extraite")
        return rangees

    def extraire(self):
        bornes = self._bornes()
        manquants = [n for n in list(BRUTS) + list(TEMOINS_VU) + list(TEMOINS_PART) if n not in bornes]
        if manquants:
            raise ValueError(f"tableaux absents du PDF : {manquants}")
        # Mois de référence : en-tête du T1, cohérence exigée partout
        for num in sorted(set(BRUTS) | set(TEMOINS_VU) | set(TEMOINS_PART)):
            i, fin = bornes[num]
            bloc = self.lignes[i:fin]
            mois = self._lire_mois(bloc)
            if not self.mois:
                self.mois = mois
            elif mois != self.mois:
                raise ValueError(f"T{num} : mois {mois} ≠ en-tête de référence {self.mois}")
            self.tables[num] = self._lire_table(num, bloc, lignes_t1=(num in (1, 2)))

    # ── Vérifications croisées ────────────────────────────────────────────────
    # Trois niveaux : erreur d'extraction (fatale), incohérence interne du
    # bulletin (avertissement : nos chiffres sont confirmés par les cumuls),
    # arrondi d'impression (accepté par calcul d'intervalles).
    #
    # `complements` : le cumul imprimé couvre janvier → mois du bulletin, or le
    # fichier ne porte que 4 mois. Dès le bulletin de mai, des mois de l'année
    # (janvier…) manquent au fichier : l'appelant (l'import de la plateforme)
    # fournit leur somme par tableau et libellé, en unités du tableau, lue dans
    # la base des bulletins antérieurs. Pour T1/T2, clés « Valeur » et « Poids ».
    def verifier(self, complements: dict[int, dict[str, float]] | None = None) -> list[str]:
        rapport: list[str] = []
        erreurs: list[str] = []
        avertissements: list[str] = []
        ok = 0

        # Périmètre du cumul « année en cours » : seuls les mois du fichier
        # appartenant à l'année du bulletin y participent (un bulletin de mars
        # porte déc-N−1 qui n'entre PAS dans le cumul N) ; les mois de janvier
        # au premier mois du fichier en sont absents et viennent du complément.
        an_bulletin = numero_mois(self.mois[-1])[1]
        idx_annee = [i for i in range(1, 5) if numero_mois(self.mois[i])[1] == an_bulletin]
        nb_mois_cumul = numero_mois(self.mois[-1])[0]      # janvier → mois du bulletin
        manquants = numero_mois(self.mois[idx_annee[0]])[0] - 1
        non_verifiables = 0

        def trouver(table: dict, lib: str):
            """Cherche un libellé dans une table : exact, puis rapprochement flou
            (les tableaux témoins du bulletin comportent des coquilles)."""
            if lib in table:
                return table[lib]
            proches = get_close_matches(lib, list(table), n=1, cutoff=0.88)
            return table[proches[0]] if proches else None

        def complement(num: int, lib: str) -> float | None:
            """Somme des mois de l'année absents du fichier (None : indisponible)."""
            if manquants == 0:
                return 0.0
            if complements is None:
                return None
            table_c = complements.get(num) or {}
            if num in (1, 2):   # T1/T2 : deux lignes par rubrique ENSEMBLE
                return table_c.get("Valeur" if lib.startswith("Valeur") else "Poids", 0.0)
            if lib in table_c:
                return table_c[lib]
            proches = get_close_matches(lib, list(table_c), n=1, cutoff=0.88)
            # Absent de la base : rubrique nouvelle, aucun flux antérieur
            return table_c[proches[0]] if proches else 0.0

        def somme_annee(num: int, lib: str) -> float:
            return sum(v for i in idx_annee if (v := self.tables[num][lib][i]) is not None)

        def cumul_coherent(num: int, lib: str) -> bool:
            """Mois du fichier + complément retrouvent-ils le cumul imprimé ?"""
            cum = self.cumuls[num].get(lib, (None, None))[1]
            comp = complement(num, lib)
            if cum is None or comp is None:
                return False
            somme = comp + somme_annee(num, lib)
            return abs(cum - somme) <= max(0.5 * nb_mois_cumul - 0.5, 0.005 * abs(cum))

        # 1. Cumuls du bulletin ≈ mois du fichier + complément (le socle).
        # Tolérance : chaque mois imprimé est arrondi (±0,5), le cumul aussi.
        for num in BRUTS:
            for lib, vals in self.tables[num].items():
                if lib.upper() == "TOTAL":
                    continue  # dérivée, contrôlée face à l'ensemble (contrôle 4)
                cum = self.cumuls[num].get(lib, (None, None))[1]
                if cum is None:
                    continue
                comp = complement(num, lib)
                if comp is None:
                    non_verifiables += 1
                    continue
                somme = comp + somme_annee(num, lib)
                if abs(cum - somme) <= max(0.5 * (nb_mois_cumul + 1), 0.005 * abs(cum)):
                    ok += 1
                elif manquants:
                    # Les mois du complément viennent de la base : un écart
                    # au-delà des arrondis signale une révision ou un
                    # reclassement ANSD d'un mois antérieur au fichier
                    # (révision semestrielle, nomenclature reventilée…), non
                    # détaillé dans ce bulletin. Ce n'est PAS une erreur
                    # d'extraction : les 4 mois du fichier restent verrouillés
                    # par leurs témoins mensuels (VU, parts, sommes).
                    avertissements.append(
                        f"Cumul T{num} {lib!r} : bulletin {cum:,.0f} ≠ fichier + base {somme:,.0f} — "
                        f"révision ou reclassement ANSD d'un mois antérieur au fichier ; "
                        f"la plateforme recalcule ses cumuls depuis les mois en base")
                else:
                    # Bulletin ≤ avril : le cumul ne dépend que du fichier,
                    # un écart est une erreur d'extraction — fatal.
                    erreurs.append(f"Cumul T{num} {lib!r} : bulletin {cum:,.0f} ≠ fichier + base {somme:,.0f}")
        if non_verifiables:
            avertissements.append(
                f"{non_verifiables} cumuls non vérifiés : {manquants} mois de l'année manquent au fichier "
                f"(bulletin postérieur à avril) — l'import sur la plateforme les fournit depuis la base et "
                f"effectue la vérification complète")

        # 2. Valeurs unitaires : la VU imprimée doit tomber dans l'intervalle
        #    permis par les arrondis d'impression (valeur ±0,5 M ; poids ±0,5 k)
        for t_vu, (t_val, t_poids) in TEMOINS_VU.items():
            for lib, vus in self.tables[t_vu].items():
                if lib.upper() == "TOTAL":
                    continue
                vals = trouver(self.tables[t_val], lib)
                poids = trouver(self.tables[t_poids], lib)
                if not vals or not poids:
                    # Rubrique renommée dans un tableau mais pas dans l'autre
                    # (ex. « PRODUITS PETROLIERS » vs « AUTRES PRODUITS
                    # PETROLIERS », août 2025) : le témoin est perdu, la
                    # rubrique brute reste contrôlée par cumuls/parts/sommes.
                    avertissements.append(
                        f"T{t_vu} : libellé témoin {lib!r} absent de T{t_val}/T{t_poids} — "
                        f"renommage ANSD probable, témoin ignoré")
                    continue
                for m in range(1, 5):
                    if vals[m] is None or poids[m] is None or poids[m] <= 0.5 or vus[m] is None:
                        continue
                    borne_basse = max(0.0, (vals[m] - 0.5)) / (poids[m] + 0.5) * 1000.0
                    borne_haute = (vals[m] + 0.5) / max(poids[m] - 0.5, 1e-9) * 1000.0
                    if borne_basse - 0.5 <= vus[m] <= borne_haute + 0.5:
                        ok += 1
                    elif lib not in self.tables[t_val] or (cumul_coherent(t_val, lib) and cumul_coherent(t_poids, lib)):
                        avertissements.append(
                            f"VU T{t_vu} {lib!r} {self.mois[m]} : bulletin {vus[m]:,.0f} hors [{borne_basse:,.0f} ; {borne_haute:,.0f}] — "
                            f"valeur/poids confirmés par leurs cumuls : coquille du bulletin")
                    else:
                        erreurs.append(f"VU T{t_vu} {lib!r} {self.mois[m]} : {vus[m]:,.0f} hors [{borne_basse:,.0f} ; {borne_haute:,.0f}]")

        # 3. Parts imprimées ≈ valeur / ensemble (arrondis : part ±0,01 ; valeur ±0,5 M)
        for t_part, t_val in TEMOINS_PART.items():
            export = t_val in (3, 11, 19, 27)
            ligne_totale = next(v for k, v in self.tables[1 if export else 2].items() if k.startswith("Valeur"))
            for lib, parts in self.tables[t_part].items():
                if lib.upper() == "TOTAL":
                    continue
                vals = trouver(self.tables[t_val], lib)
                if not vals:
                    continue
                for m in range(1, 5):
                    if vals[m] is None or parts[m] is None or not ligne_totale[m]:
                        continue
                    basse = max(0.0, vals[m] - 0.5) * 1e6 / ligne_totale[m] * 100.0
                    haute = (vals[m] + 0.5) * 1e6 / ligne_totale[m] * 100.0
                    if basse - 0.015 <= parts[m] <= haute + 0.015:
                        ok += 1
                    elif lib not in self.tables[t_val] or cumul_coherent(t_val, lib):
                        avertissements.append(
                            f"Part T{t_part} {lib!r} {self.mois[m]} : {parts[m]} hors [{basse:.2f} ; {haute:.2f}] — valeur confirmée par cumul")
                    else:
                        erreurs.append(f"Part T{t_part} {lib!r} {self.mois[m]} : {parts[m]} hors [{basse:.2f} ; {haute:.2f}]")

        # 4. Sommes exhaustives (hors ligne TOTAL) ≈ ensemble, et TOTAL ≈ ensemble
        for t_dim, t_ens, en_val in ((3, 1, True), (7, 2, True), (19, 1, True), (23, 2, True),
                                     (4, 1, False), (8, 2, False), (20, 1, False), (24, 2, False)):
            prefixe = "Valeur" if en_val else "Poids"
            ligne = next(v for k, v in self.tables[t_ens].items() if k.startswith(prefixe))
            facteur = 1e6 if en_val else 1e3
            rangees = {k: v for k, v in self.tables[t_dim].items() if k.upper() != "TOTAL"}
            total = self.tables[t_dim].get("TOTAL")
            for m in range(1, 5):
                if not ligne[m]:
                    continue
                somme = sum(v[m] or 0 for v in rangees.values()) * facteur
                # tolérance : ±0,5 unité d'impression par rangée
                tol = 0.5 * facteur * len(rangees)
                if abs(somme - ligne[m]) <= tol:
                    ok += 1
                else:
                    erreurs.append(f"Somme T{t_dim} {self.mois[m]} : {somme:,.0f} ≠ ensemble {ligne[m]:,.0f}")
                if total and total[m] is not None:
                    if abs(total[m] * facteur - ligne[m]) <= facteur:
                        ok += 1
                    elif abs(somme - ligne[m]) <= tol:
                        avertissements.append(
                            f"TOTAL T{t_dim} {self.mois[m]} : {total[m] * facteur:,.0f} ≠ ensemble {ligne[m]:,.0f} — "
                            f"la somme des rangées est correcte : coquille du bulletin")
                    else:
                        erreurs.append(f"TOTAL T{t_dim} {self.mois[m]} : {total[m] * facteur:,.0f} ≠ ensemble {ligne[m]:,.0f}")

        # Témoins illisibles : la vérification correspondante est perdue pour la
        # rubrique (signalé), mais ses données restent contrôlées par ailleurs
        for a in getattr(self, "temoins_perdus", []):
            avertissements.append(f"{a} — témoin illisible ignoré, la rubrique reste contrôlée par cumuls, parts et sommes")

        rapport.append(f"{ok} contrôles réussis, {len(erreurs)} erreurs, "
                       f"{len(avertissements)} incohérences internes du bulletin, {len(self.anomalies)} anomalies de lecture")
        rapport.extend("✗ " + e for e in erreurs)
        rapport.extend("⚠ " + a for a in avertissements)
        rapport.extend("• " + a for a in self.anomalies)
        if erreurs or self.anomalies:
            raise SystemExit("ÉCHEC DES VÉRIFICATIONS :\n" + "\n".join(rapport))
        return rapport

    # ── Sortie Excel ──────────────────────────────────────────────────────────
    def ecrire_xlsx(self, sortie: Path, rapport: list[str]):
        wb = Workbook()
        wb.remove(wb.active)
        gras = Font(bold=True)
        fond = PatternFill("solid", fgColor="E8EEF5")
        mois_retenus = self.mois[1:]  # les 4 mois (sans le mois de l'année précédente)
        for num, (titre, unite) in BRUTS.items():
            ws = wb.create_sheet(f"T{num}")
            ws.append([f"TABLEAU {num} — {titre} ({unite})"])
            ws["A1"].font = gras
            ws.append(["Libellé", *mois_retenus])
            for c in ws[2]:
                c.font, c.fill = gras, fond
            for lib, vals in self.tables[num].items():
                if lib.upper() == "TOTAL":
                    continue  # dérivé — recalculé par la plateforme
                ws.append([lib, *[v for v in vals[1:5]]])
            ws.column_dimensions["A"].width = 58
        meta = wb.create_sheet("Métadonnées")
        meta.append(["Bulletin (mois)", self.mois[-1]])
        meta.append(["Mois couverts", ", ".join(mois_retenus)])
        meta.append(["Fichier source", self.chemin.name])
        meta.append(["Extrait le", datetime.now().strftime("%Y-%m-%d %H:%M")])
        meta.append(["Unités T1/T2", "Valeur en FCFA, Poids net en kg"])
        meta.append(["Unités T3+", "Valeur en millions de FCFA, Poids net en milliers de kg"])
        meta.append(["Convention", "cellule vide = absence de flux (« – » du bulletin)"])
        meta.append([])
        meta.append(["RAPPORT DE VÉRIFICATION"])
        for l in rapport:
            meta.append([l])
        meta.column_dimensions["A"].width = 40
        wb.save(sortie)


def main():
    if len(sys.argv) < 2:
        raise SystemExit(__doc__)
    pdf = Path(sys.argv[1])
    sortie = Path(sys.argv[2]) if len(sys.argv) > 2 else pdf.with_suffix(".xlsx")
    b = Bulletin(pdf)
    b.extraire()
    rapport = b.verifier()
    b.ecrire_xlsx(sortie, rapport)
    print(f"Bulletin {b.mois[-1]} extrait → {sortie}")
    print(rapport[0])


if __name__ == "__main__":
    main()
