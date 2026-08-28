#!/usr/bin/env python3
"""Dérive les CSV de la classification fDi Markets depuis les deux classeurs source.

    python backend/scripts/fdi/generer_csv.py

Lit `source/fdi_classification_fr.xlsx` (l'arbre secteur → sous-secteur, en
français) et `source/fdi_correspondance_en_fr.xlsx` (les trois tables de
correspondance anglais ↔ français), et écrit trois CSV à côté :

    fdi_secteurs.csv · fdi_sous_secteurs.csv · fdi_business_activites.csv

Ce sont les CSV, et non les classeurs, qui font foi pour l'import : ils se
relisent dans une revue de code, se comparent d'une version à l'autre, et
n'imposent pas de dépendance Excel au conteneur backend. Les classeurs restent
versionnés à côté pour que la dérivation soit rejouable et vérifiable.

Le script est *pur* : mêmes classeurs, mêmes CSV — aucun horodatage, aucun
identifiant de base. Il échoue plutôt que d'écrire un CSV douteux (voir les
sept contrôles en fin de fichier).
"""
from __future__ import annotations

import csv
import re
import sys
import unicodedata
from collections import Counter
from pathlib import Path

try:
    import openpyxl
except ModuleNotFoundError:  # pragma: no cover — dépendance de développement
    sys.exit("openpyxl requis :  pip install openpyxl")

# Les dérivations (slug, clé d'appariement) vivent dans le service : l'import
# des fichiers et la saisie à l'écran doivent produire les mêmes formes, et deux
# copies d'une même règle finissent toujours par diverger.
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from app.services.fdi_classification import cle_de, sans_parenthese, slug  # noqa: E402

DOSSIER = Path(__file__).resolve().parent
SOURCE = DOSSIER / "source"
CLASSIFICATION = SOURCE / "fdi_classification_fr.xlsx"
CORRESPONDANCE = SOURCE / "fdi_correspondance_en_fr.xlsx"

# Longueur maximale de la partie « libellé » d'un code de sous-secteur. Les
# libellés fDi vont jusqu'à 100 caractères (« Media streaming distribution
# services, social networks… ») ; tronquer garde des codes maniables, et le
# contrôle d'unicité en fin de script refuse la troncature du jour où deux
# libellés d'un même secteur partageraient leurs 40 premiers caractères.
CAP_CODE = 40


def lire_classification() -> tuple[list[tuple[str, str]], list[str]]:
    """L'arbre FR : couples (secteur, sous-secteur), et la liste des activités.

    La colonne « Secteur » n'est renseignée que sur la première ligne de chaque
    groupe — mise en forme de tableur, pas donnée manquante : on la propage
    vers le bas.
    """
    wb = openpyxl.load_workbook(CLASSIFICATION, read_only=True, data_only=True)
    couples: list[tuple[str, str]] = []
    secteur: str | None = None
    for i, ligne in enumerate(wb["Secteurs et sous-secteurs"].iter_rows(values_only=True), start=1):
        if i <= 3:  # titre, ligne vide, en-têtes
            continue
        gauche, droite = ligne[0], ligne[1]
        if gauche and str(gauche).strip():
            secteur = str(gauche).strip()
        if droite and str(droite).strip():
            if secteur is None:
                sys.exit(f"ligne {i} : sous-secteur sans secteur parent")
            couples.append((secteur, str(droite).strip()))

    activites = [
        str(ligne[1]).strip()
        for i, ligne in enumerate(wb["Activités économiques"].iter_rows(values_only=True), start=1)
        if i > 3 and ligne[1] and str(ligne[1]).strip()
    ]
    return couples, activites


def lire_correspondance(feuille: str) -> dict[str, str]:
    """Une feuille de correspondance, indexée par le libellé français."""
    wb = openpyxl.load_workbook(CORRESPONDANCE, read_only=True, data_only=True)
    table: dict[str, str] = {}
    for i, ligne in enumerate(wb[feuille].iter_rows(values_only=True), start=1):
        if i == 1 or not ligne[0] or not ligne[1]:
            continue
        en, fr = str(ligne[0]).strip(), str(ligne[1]).strip()
        if fr in table and table[fr] != en:
            sys.exit(f"{feuille} : « {fr} » traduit deux fois ({table[fr]} / {en})")
        table[fr] = en
    return table


def main() -> int:
    couples, activites_fr = lire_classification()
    trad_secteur = lire_correspondance("Secteurs")
    trad_sous = lire_correspondance("Sous-secteurs")
    trad_activite = lire_correspondance("Business activities")

    # ── Secteurs, dans l'ordre du classeur ────────────────────────────────────
    secteurs_fr: list[str] = list(dict.fromkeys(s for s, _ in couples))
    manquants = [s for s in secteurs_fr if s not in trad_secteur]
    if manquants:
        sys.exit(f"secteurs sans traduction anglaise : {manquants}")
    secteurs = [
        {"code": slug(trad_secteur[fr]), "libelle_en": trad_secteur[fr], "libelle_fr": fr,
         "ordre": i + 1}
        for i, fr in enumerate(secteurs_fr)
    ]
    code_de_secteur = {s["libelle_fr"]: s["code"] for s in secteurs}

    # ── Sous-secteurs ─────────────────────────────────────────────────────────
    manquants = [ss for _, ss in couples if ss not in trad_sous]
    if manquants:
        sys.exit(f"sous-secteurs sans traduction anglaise : {manquants[:5]}")
    sous_secteurs = []
    for i, (secteur_fr, fr) in enumerate(couples):
        en = trad_sous[fr]
        base_en = sans_parenthese(en)
        code = f"{code_de_secteur[secteur_fr]}__{slug(base_en)[:CAP_CODE].rstrip('_')}"
        sous_secteurs.append({
            "code": code,
            "secteur_code": code_de_secteur[secteur_fr],
            # Verbatim : c'est ce libellé, parenthèse comprise, que porteront
            # les exports de projets fDi. Toute « correction » ici casserait
            # l'appariement — y compris la coquille « infrastucture », qui est
            # celle de la source et doit le rester.
            "libelle_en": en,
            "libelle_fr": fr,
            "libelle_en_base": base_en,
            "cle_appariement": cle_de(base_en),
            "ordre": i + 1,
        })

    # ── Activités économiques (sans lien avec l'arbre sectoriel) ──────────────
    manquants = [a for a in activites_fr if a not in trad_activite]
    if manquants:
        sys.exit(f"activités sans traduction anglaise : {manquants}")
    activites = [
        {"code": slug(trad_activite[fr]), "libelle_en": trad_activite[fr], "libelle_fr": fr,
         "cle_appariement": cle_de(trad_activite[fr]), "ordre": i + 1}
        for i, fr in enumerate(activites_fr)
    ]

    # ── Sept contrôles avant écriture ─────────────────────────────────────────
    def unique(nom: str, valeurs: list[str]) -> None:
        doublons = {v: n for v, n in Counter(valeurs).items() if n > 1}
        if doublons:
            sys.exit(f"{nom} : valeurs en double {list(doublons)[:5]}")

    unique("code secteur", [s["code"] for s in secteurs])
    unique("libellé EN secteur", [s["libelle_en"] for s in secteurs])
    unique("code sous-secteur", [s["code"] for s in sous_secteurs])
    unique("libellé EN sous-secteur", [s["libelle_en"] for s in sous_secteurs])
    unique("code activité", [a["code"] for a in activites])
    # La clé d'appariement n'est unique QUE dans son secteur : c'est la raison
    # d'être de la parenthèse de fDi, et donc de la clé composite en base.
    for s in secteurs:
        cles = [x["cle_appariement"] for x in sous_secteurs if x["secteur_code"] == s["code"]]
        unique(f"clé d'appariement dans « {s['libelle_en']} »", cles)
    orphelins = [x["code"] for x in sous_secteurs if x["secteur_code"] not in code_de_secteur.values()]
    if orphelins:
        sys.exit(f"sous-secteurs rattachés à un secteur inconnu : {orphelins[:5]}")

    # ── Écriture ──────────────────────────────────────────────────────────────
    def ecrire(nom: str, lignes: list[dict]) -> None:
        chemin = DOSSIER / nom
        with chemin.open("w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=list(lignes[0].keys()))
            w.writeheader()
            w.writerows(lignes)
        print(f"  {nom:<32} {len(lignes):>4} lignes")

    ecrire("fdi_secteurs.csv", secteurs)
    ecrire("fdi_sous_secteurs.csv", sous_secteurs)
    ecrire("fdi_business_activites.csv", activites)

    partages = Counter(x["cle_appariement"] for x in sous_secteurs)
    multi = sum(1 for n in partages.values() if n > 1)
    print(f"  → {multi} libellés de sous-secteur partagés par plusieurs secteurs "
          f"(« Other » en tête) : l'appariement exige donc toujours le secteur.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
